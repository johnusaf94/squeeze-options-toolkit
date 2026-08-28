"""
bogle_analyzer.py
==================
John Bogle's framework applied as concrete portfolio analysis.

Three pillars:
1. Don't Fight the Last War — is past performance reliable or is this chasing?
2. Reversion to the Mean — is the stock/fund stretched vs its history? Buy timing signal.
3. Portfolio Diversification — sector allocation by value AND income, impact of adding new position.

Requires: pip install yfinance requests openpyxl
"""

import yfinance as yf
import numpy as np
from dataclasses import dataclass, field
from typing import Optional
import os

# ─────────────────────────────────────────────
# BOGLE THRESHOLDS
# ─────────────────────────────────────────────

# Reversion windows (trading days)
RTM_SHORT_WINDOW  = 63    # ~3 months
RTM_MEDIUM_WINDOW = 252   # ~1 year
RTM_LONG_WINDOW   = 756   # ~3 years

# How stretched is "stretched"
RTM_OVERBOUGHT_Z  = 1.5   # z-score above mean = overbought
RTM_OVERSOLD_Z    = -1.5  # z-score below mean = oversold

# Don't Fight the Last War — lookback
DFTLW_PERIOD_YRS  = 5     # years of return to evaluate
DFTLW_VOLATILITY_THRESHOLD = 0.25  # annualized vol > 25% = high vol regime

# Concentration thresholds (Bogle's diversification warnings)
SECTOR_WARN_PCT   = 0.30   # single sector > 30% of portfolio = concentration warning
SECTOR_CRITICAL   = 0.50   # single sector > 50% = critical
INCOME_WARN_PCT   = 0.40   # single ticker > 40% of income = income concentration warning

# ─────────────────────────────────────────────
# DATA STRUCTURES
# ─────────────────────────────────────────────

@dataclass
class LastWarAnalysis:
    """Past Performance Reliability — is recent outperformance structural or a regime fluke?"""
    ticker:                 str = ""
    period_years:           int = DFTLW_PERIOD_YRS
    annualized_return_pct:  Optional[float] = None   # actual historical return
    annualized_volatility:  Optional[float] = None   # historical vol
    sharpe_ratio:           Optional[float] = None   # return/vol (risk-adjusted)
    max_drawdown:           Optional[float] = None   # worst peak-to-trough
    beta_to_spy:            Optional[float] = None   # market sensitivity
    return_vs_spy:          Optional[float] = None   # alpha vs S&P 500

    # The key Bogle question: is this performance structural or cyclical?
    performance_regime:     str = "UNKNOWN"          # STRUCTURAL / CYCLICAL / MIXED
    regime_reasoning:       str = ""
    last_war_risk:          str = "UNKNOWN"          # LOW / MODERATE / HIGH
    last_war_warning:       str = ""


@dataclass
class ReversionSignal:
    """Reversion to the Mean — where is the stock relative to its own history?"""
    ticker:                 str = ""

    # Price vs moving averages
    price_current:          Optional[float] = None
    ma_50d:                 Optional[float] = None
    ma_200d:                Optional[float] = None
    pct_above_ma50:         Optional[float] = None
    pct_above_ma200:        Optional[float] = None

    # Z-score: how many std devs from mean price?
    z_score_1yr:            Optional[float] = None   # vs 1yr mean
    z_score_3yr:            Optional[float] = None   # vs 3yr mean

    # P/E vs historical range
    pe_current:             Optional[float] = None
    pe_5yr_avg:             Optional[float] = None
    pe_stretch:             Optional[float] = None   # % above/below 5yr avg P/E

    # RSI (momentum proxy)
    rsi_14:                 Optional[float] = None

    # Buy timing signal
    timing_signal:          str = "UNKNOWN"   # BEST / GOOD / MEDIUM / BAD
    timing_score:           int = 0           # 0-10
    timing_reasoning:       str = ""


@dataclass
class HoldingAllocation:
    ticker:         str = ""
    company_name:   str = ""
    sector:         str = ""
    market_value:   float = 0.0
    annual_income:  float = 0.0
    dividend_yield: float = 0.0
    account:        str = ""
    shares:         float = 0.0


@dataclass
class DiversificationAnalysis:
    """Full portfolio diversification breakdown + impact of adding new position."""

    # Current portfolio
    holdings:               list = field(default_factory=list)   # list of HoldingAllocation
    total_value:            float = 0.0
    total_annual_income:    float = 0.0

    # Sector breakdowns
    sector_by_value:        dict = field(default_factory=dict)   # sector -> % of portfolio value
    sector_by_income:       dict = field(default_factory=dict)   # sector -> % of portfolio income
    sector_value_dollars:   dict = field(default_factory=dict)
    sector_income_dollars:  dict = field(default_factory=dict)

    # Concentration warnings
    value_warnings:         list = field(default_factory=list)
    income_warnings:        list = field(default_factory=list)
    herfindahl_index:       Optional[float] = None   # 0=perfect diversification, 1=total concentration
    diversification_grade:  str = "N/A"              # A/B/C/D/F

    # Impact of adding new ticker
    new_ticker:             str = ""
    new_ticker_sector:      str = ""
    new_shares_assumed:     int = 0
    new_position_value:     float = 0.0
    new_annual_income:      float = 0.0

    after_sector_by_value:  dict = field(default_factory=dict)
    after_sector_by_income: dict = field(default_factory=dict)
    diversification_impact: str = "NEUTRAL"   # IMPROVES / NEUTRAL / HURTS / HURTS_SIGNIFICANTLY
    impact_reasoning:       str = ""


@dataclass
class BogleAnalysis:
    ticker:             str = ""
    company_name:       str = ""
    sector:             str = ""
    last_war:           LastWarAnalysis = field(default_factory=LastWarAnalysis)
    reversion:          ReversionSignal = field(default_factory=ReversionSignal)
    diversification:    DiversificationAnalysis = field(default_factory=DiversificationAnalysis)
    errors:             list = field(default_factory=list)


# ─────────────────────────────────────────────
# PILLAR 1: DON'T FIGHT THE LAST WAR
# ─────────────────────────────────────────────

def analyze_last_war(ticker: str, info: dict) -> LastWarAnalysis:
    lw = LastWarAnalysis(ticker=ticker)  # Past Performance Reliability check

    try:
        t = yf.Ticker(ticker)
        spy = yf.Ticker("SPY")

        hist = t.history(period=f"{DFTLW_PERIOD_YRS}y", interval="1mo")
        spy_hist = spy.history(period=f"{DFTLW_PERIOD_YRS}y", interval="1mo")

        if hist.empty or len(hist) < 12:
            lw.last_war_risk = "INSUFFICIENT DATA"
            return lw

        # Monthly returns
        returns = hist["Close"].pct_change().dropna()
        spy_returns = spy_hist["Close"].pct_change().dropna()

        # Align on common dates
        common = returns.index.intersection(spy_returns.index)
        returns = returns[common]
        spy_returns = spy_returns[common]

        # Annualized return and vol
        ann_return = (1 + returns.mean()) ** 12 - 1
        ann_vol = returns.std() * (12 ** 0.5)
        lw.annualized_return_pct = float(ann_return)
        lw.annualized_volatility = float(ann_vol)

        # Sharpe (using 4.5% risk-free rate)
        rf = 0.045 / 12
        excess = returns - rf
        lw.sharpe_ratio = float(excess.mean() / excess.std() * (12 ** 0.5)) if excess.std() > 0 else None

        # Max drawdown
        cumulative = (1 + returns).cumprod()
        rolling_max = cumulative.cummax()
        drawdown = (cumulative - rolling_max) / rolling_max
        lw.max_drawdown = float(drawdown.min())

        # Beta vs SPY
        if len(spy_returns) > 5:
            cov = returns.cov(spy_returns)
            var = spy_returns.var()
            lw.beta_to_spy = float(cov / var) if var > 0 else None

        # Return vs SPY
        spy_ann = (1 + spy_returns.mean()) ** 12 - 1
        lw.return_vs_spy = float(ann_return - spy_ann)

        # ── PAST PERFORMANCE RELIABILITY ASSESSMENT ──────────────────────────────────────────────
        # Bogle's key insight: past returns in high-beta, high-vol, or sector-
        # concentrated assets are least likely to repeat. Broad market returns
        # are most likely to persist (regression toward long-run equity premium).

        risk_factors = 0
        warnings = []

        # High volatility = return likely cyclical
        if ann_vol > DFTLW_VOLATILITY_THRESHOLD:
            risk_factors += 2
            warnings.append(f"High volatility ({ann_vol:.0%} annualized) — cyclical, not structural")

        # High beta = returns driven by market regime, not company quality
        if lw.beta_to_spy and lw.beta_to_spy > 1.5:
            risk_factors += 2
            warnings.append(f"Beta {lw.beta_to_spy:.2f} — highly market-regime dependent")

        # Massive outperformance is mean-reverting
        if lw.return_vs_spy and lw.return_vs_spy > 0.10:
            risk_factors += 1
            warnings.append(f"Outperformed SPY by {lw.return_vs_spy:.1%}/yr — statistically likely to revert")

        # Severe drawdown suggests fragility
        if lw.max_drawdown and lw.max_drawdown < -0.40:
            risk_factors += 1
            warnings.append(f"Max drawdown {lw.max_drawdown:.0%} — history of severe crashes")

        # Low Sharpe = returns not worth the risk
        if lw.sharpe_ratio and lw.sharpe_ratio < 0.5:
            risk_factors += 1
            warnings.append(f"Sharpe {lw.sharpe_ratio:.2f} — poor risk-adjusted return")

        if risk_factors == 0:
            lw.performance_regime = "STRUCTURAL"
            lw.last_war_risk = "LOW"
            lw.regime_reasoning = "Returns appear driven by fundamental quality, not regime or cycle."
        elif risk_factors <= 2:
            lw.performance_regime = "MIXED"
            lw.last_war_risk = "MODERATE"
            lw.regime_reasoning = "Some cyclical factors present — future may not mirror past exactly."
        else:
            lw.performance_regime = "CYCLICAL"
            lw.last_war_risk = "HIGH"
            lw.regime_reasoning = "Past performance heavily regime-driven — Bogle would warn against extrapolating."

        lw.last_war_warning = " | ".join(warnings) if warnings else "No significant last-war risks detected."

    except Exception as e:
        lw.last_war_risk = "ERROR"
        lw.last_war_warning = str(e)

    return lw


# ─────────────────────────────────────────────
# PILLAR 2: REVERSION TO THE MEAN
# ─────────────────────────────────────────────

def calculate_rsi(prices, period=14):
    delta = prices.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.rolling(window=period).mean()
    avg_loss = loss.rolling(window=period).mean()
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi.iloc[-1] if not rsi.empty else None


def analyze_reversion(ticker: str, info: dict) -> ReversionSignal:
    rv = ReversionSignal(ticker=ticker)

    try:
        t = yf.Ticker(ticker)
        hist_3yr = t.history(period="3y", interval="1d")

        if hist_3yr.empty or len(hist_3yr) < 50:
            rv.timing_signal = "INSUFFICIENT DATA"
            return rv

        prices = hist_3yr["Close"]
        rv.price_current = float(prices.iloc[-1])

        # Moving averages
        if len(prices) >= 50:
            rv.ma_50d = float(prices.rolling(50).mean().iloc[-1])
            rv.pct_above_ma50 = (rv.price_current - rv.ma_50d) / rv.ma_50d

        if len(prices) >= 200:
            rv.ma_200d = float(prices.rolling(200).mean().iloc[-1])
            rv.pct_above_ma200 = (rv.price_current - rv.ma_200d) / rv.ma_200d

        # Z-scores (price vs its own history)
        hist_1yr = prices.iloc[-252:] if len(prices) >= 252 else prices
        rv.z_score_1yr = float((rv.price_current - hist_1yr.mean()) / hist_1yr.std()) if hist_1yr.std() > 0 else None
        rv.z_score_3yr = float((rv.price_current - prices.mean()) / prices.std()) if prices.std() > 0 else None

        # P/E vs 5yr average (approximate — yfinance doesn't give historical P/E directly)
        rv.pe_current = info.get("trailingPE")
        # Use forward P/E 5yr avg if available, else approximate from earnings yield history
        rv.pe_5yr_avg = info.get("fiveYearAverageReturn")  # proxy
        if rv.pe_current and rv.pe_5yr_avg:
            rv.pe_stretch = (rv.pe_current - rv.pe_5yr_avg) / rv.pe_5yr_avg

        # RSI
        rv.rsi_14 = calculate_rsi(prices, 14)
        if rv.rsi_14:
            rv.rsi_14 = float(rv.rsi_14)

        # ── TIMING SCORE ─────────────────────────────────────────────────────
        # 0 = worst time to buy (most overbought)
        # 10 = best time to buy (most oversold)
        score = 5   # neutral start
        reasons = []

        # Z-score contribution (most important)
        if rv.z_score_3yr is not None:
            if rv.z_score_3yr <= -1.5:
                score += 3
                reasons.append(f"Price {abs(rv.z_score_3yr):.1f} std devs BELOW 3yr mean — historically cheap")
            elif rv.z_score_3yr <= -0.5:
                score += 1
                reasons.append(f"Price slightly below 3yr mean — mild tailwind for mean reversion")
            elif rv.z_score_3yr >= 1.5:
                score -= 3
                reasons.append(f"Price {rv.z_score_3yr:.1f} std devs ABOVE 3yr mean — stretched, expect reversion")
            elif rv.z_score_3yr >= 0.5:
                score -= 1
                reasons.append(f"Price above 3yr mean — slight headwind from reversion pressure")

        # MA contribution
        if rv.pct_above_ma200 is not None:
            if rv.pct_above_ma200 < -0.10:
                score += 2
                reasons.append(f"{abs(rv.pct_above_ma200):.0%} below 200d MA — significant discount to trend")
            elif rv.pct_above_ma200 > 0.20:
                score -= 2
                reasons.append(f"{rv.pct_above_ma200:.0%} above 200d MA — extended above long-term trend")

        # RSI contribution
        if rv.rsi_14 is not None:
            if rv.rsi_14 < 35:
                score += 2
                reasons.append(f"RSI {rv.rsi_14:.0f} — oversold momentum signal")
            elif rv.rsi_14 > 70:
                score -= 2
                reasons.append(f"RSI {rv.rsi_14:.0f} — overbought momentum signal")

        # Clamp score
        score = max(0, min(10, score))
        rv.timing_score = score

        if score >= 8:
            rv.timing_signal = "BEST"
        elif score >= 6:
            rv.timing_signal = "GOOD"
        elif score >= 4:
            rv.timing_signal = "MEDIUM"
        else:
            rv.timing_signal = "BAD"

        rv.timing_reasoning = " | ".join(reasons) if reasons else "Price near historical mean — neutral timing."

    except Exception as e:
        rv.timing_signal = "ERROR"
        rv.timing_reasoning = str(e)

    return rv


# ─────────────────────────────────────────────
# PILLAR 3: PORTFOLIO DIVERSIFICATION
# ─────────────────────────────────────────────

def load_portfolio_holdings(portfolio_path: str = "portfolio.xlsx") -> list:
    """Load current holdings from the Excel file."""
    holdings = []

    if not os.path.exists(portfolio_path):
        return holdings

    try:
        import openpyxl
        wb = openpyxl.load_workbook(portfolio_path, data_only=True)
        if "Holdings" not in wb.sheetnames:
            return holdings

        ws = wb["Holdings"]
        headers = []
        for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
            vals = [v for v in row]
            if i == 0:
                headers = [str(v).strip() if v else "" for v in vals]
                continue

            if not vals[0] or str(vals[0]).strip() in ("", "TOTALS"):
                continue

            ticker = str(vals[0]).strip().upper()
            account = str(vals[1]).strip() if vals[1] else ""
            shares = float(vals[2]) if vals[2] else 0
            curr_price = float(vals[4]) if vals[4] else 0
            ann_div_per_share = float(vals[9]) if len(vals) > 9 and vals[9] else 0
            goal = str(vals[12]).strip() if len(vals) > 12 and vals[12] else ""

            market_value = shares * curr_price
            annual_income = shares * ann_div_per_share

            if market_value > 0:
                holdings.append(HoldingAllocation(
                    ticker=ticker,
                    account=account,
                    shares=shares,
                    market_value=market_value,
                    annual_income=annual_income,
                    dividend_yield=ann_div_per_share / curr_price if curr_price > 0 else 0,
                ))

    except Exception:
        pass

    return holdings


def enrich_holdings_with_sectors(holdings: list) -> list:
    """Add sector data to each holding via yfinance."""
    enriched = []
    for h in holdings:
        try:
            t = yf.Ticker(h.ticker)
            info = t.info
            h.company_name = info.get("longName", h.ticker)
            h.sector = info.get("sector") or _guess_sector(h.ticker)
        except Exception:
            h.sector = _guess_sector(h.ticker)
            h.company_name = h.ticker
        enriched.append(h)
    return enriched


def _guess_sector(ticker: str) -> str:
    """Fallback sector map for common tickers and ETFs."""
    known = {
        "SCHD": "Diversified ETF", "VOO": "Diversified ETF", "VTI": "Diversified ETF",
        "SPY": "Diversified ETF", "QQQ": "Technology ETF",
        "O":    "Real Estate", "STAG": "Real Estate", "REALTY": "Real Estate",
        "MSFT": "Technology", "AAPL": "Technology", "GOOGL": "Technology",
        "AMZN": "Consumer Discretionary", "TSLA": "Consumer Discretionary",
        "JNJ":  "Healthcare", "PFE": "Healthcare", "ABBV": "Healthcare",
        "KO":   "Consumer Staples", "PG": "Consumer Staples", "GIS": "Consumer Staples",
        "VZ":   "Communication Services", "T": "Communication Services",
        "XOM":  "Energy", "CVX": "Energy",
        "JPM":  "Financials", "BAC": "Financials",
    }
    return known.get(ticker.upper(), "Unknown")


def herfindahl_index(weights: dict) -> float:
    """
    Herfindahl-Hirschman Index — measure of concentration.
    0 = perfectly diversified, 1 = single holding.
    """
    return sum(w ** 2 for w in weights.values())


def diversification_grade(hhi: float) -> str:
    if hhi < 0.10: return "A — Well Diversified"
    if hhi < 0.18: return "B — Adequately Diversified"
    if hhi < 0.25: return "C — Moderate Concentration"
    if hhi < 0.40: return "D — Concentrated"
    return "F — Highly Concentrated"


def analyze_diversification(ticker: str, new_ticker_info: dict,
                             portfolio_path: str = "portfolio.xlsx",
                             new_shares: int = 10) -> DiversificationAnalysis:
    da = DiversificationAnalysis(new_ticker=ticker)

    # Load and enrich current holdings
    holdings = load_portfolio_holdings(portfolio_path)
    if not holdings:
        da.impact_reasoning = "No portfolio data found — place portfolio.xlsx in the same folder."
        return da

    holdings = enrich_holdings_with_sectors(holdings)
    da.holdings = holdings
    da.total_value = sum(h.market_value for h in holdings)
    da.total_annual_income = sum(h.annual_income for h in holdings)

    # Build sector maps
    for h in holdings:
        s = h.sector or "Unknown"
        da.sector_value_dollars[s]  = da.sector_value_dollars.get(s, 0) + h.market_value
        da.sector_income_dollars[s] = da.sector_income_dollars.get(s, 0) + h.annual_income

    if da.total_value > 0:
        da.sector_by_value = {s: v / da.total_value for s, v in da.sector_value_dollars.items()}
    if da.total_annual_income > 0:
        da.sector_by_income = {s: v / da.total_annual_income for s, v in da.sector_income_dollars.items()}

    # Warnings
    for s, pct in da.sector_by_value.items():
        if pct >= SECTOR_CRITICAL:
            da.value_warnings.append(f"🔴 CRITICAL: {s} = {pct:.0%} of portfolio value")
        elif pct >= SECTOR_WARN_PCT:
            da.value_warnings.append(f"⚠️  WARNING: {s} = {pct:.0%} of portfolio value")

    for s, pct in da.sector_by_income.items():
        if pct >= INCOME_WARN_PCT:
            da.income_warnings.append(f"⚠️  {s} = {pct:.0%} of portfolio income")

    # HHI and grade
    if da.sector_by_value:
        da.herfindahl_index = herfindahl_index(da.sector_by_value)
        da.diversification_grade = diversification_grade(da.herfindahl_index)

    # New position impact
    new_sector = new_ticker_info.get("sector") or _guess_sector(ticker)
    da.new_ticker_sector = new_sector
    new_price = new_ticker_info.get("currentPrice") or new_ticker_info.get("regularMarketPrice") or 0
    new_div_rate = new_ticker_info.get("dividendRate") or 0
    da.new_shares_assumed = new_shares
    da.new_position_value = new_price * new_shares
    da.new_annual_income = new_div_rate * new_shares

    # Recalculate after adding
    after_value = da.total_value + da.new_position_value
    after_income = da.total_annual_income + da.new_annual_income

    after_val_dollars = dict(da.sector_value_dollars)
    after_val_dollars[new_sector] = after_val_dollars.get(new_sector, 0) + da.new_position_value
    after_inc_dollars = dict(da.sector_income_dollars)
    after_inc_dollars[new_sector] = after_inc_dollars.get(new_sector, 0) + da.new_annual_income

    if after_value > 0:
        da.after_sector_by_value = {s: v / after_value for s, v in after_val_dollars.items()}
    if after_income > 0:
        da.after_sector_by_income = {s: v / after_income for s, v in after_inc_dollars.items()}

    # Assess impact
    after_hhi = herfindahl_index(da.after_sector_by_value) if da.after_sector_by_value else da.herfindahl_index
    before_hhi = da.herfindahl_index or 1.0
    delta_hhi = after_hhi - before_hhi

    existing_sector_pct = da.sector_by_value.get(new_sector, 0)

    reasons = []

    if new_sector in ("Diversified ETF",):
        da.diversification_impact = "IMPROVES"
        reasons.append(f"{ticker} is a diversified fund — inherently reduces concentration")
    elif existing_sector_pct >= SECTOR_CRITICAL:
        da.diversification_impact = "HURTS_SIGNIFICANTLY"
        reasons.append(f"{new_sector} already at {existing_sector_pct:.0%} of portfolio — adding more is dangerous concentration")
    elif existing_sector_pct >= SECTOR_WARN_PCT:
        da.diversification_impact = "HURTS"
        reasons.append(f"{new_sector} already at {existing_sector_pct:.0%} — adds to existing concentration risk")
    elif existing_sector_pct == 0:
        da.diversification_impact = "IMPROVES"
        reasons.append(f"New sector exposure — {new_sector} not currently in portfolio")
    elif delta_hhi < -0.01:
        da.diversification_impact = "IMPROVES"
        reasons.append(f"Reduces HHI concentration index by {abs(delta_hhi):.3f}")
    elif delta_hhi > 0.02:
        da.diversification_impact = "HURTS"
        reasons.append(f"Increases HHI concentration index by {delta_hhi:.3f}")
    else:
        da.diversification_impact = "NEUTRAL"
        reasons.append(f"Minimal diversification impact — {new_sector} at {existing_sector_pct:.0%} of portfolio")

    da.impact_reasoning = " | ".join(reasons)
    return da


# ─────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────

def run_bogle_analysis(ticker: str, portfolio_path: str = "portfolio.xlsx",
                       new_shares: int = 10) -> BogleAnalysis:
    analysis = BogleAnalysis(ticker=ticker.upper())

    try:
        t = yf.Ticker(ticker)
        info = t.info
        analysis.company_name = info.get("longName", ticker)
        analysis.sector = info.get("sector", "Unknown")
    except Exception as e:
        analysis.errors.append(f"Info fetch failed: {e}")
        info = {}

    # ── DATA VALIDATION GATE ──
    from data_validator import validate
    dq = validate(ticker, info, "bogle")
    analysis.errors.append(f"VALIDATION:{dq.confidence}:{dq.can_analyze}:{dq.asset_type}")
    if dq.warnings:
        for w in dq.warnings:
            analysis.errors.append(f"WARNING:{w}")
    # Bogle works on everything — ETFs are his specialty — so no hard block
    # But we still record asset type for proper framing

    analysis.last_war     = analyze_last_war(ticker, info)
    analysis.reversion    = analyze_reversion(ticker, info)
    analysis.diversification = analyze_diversification(
        ticker, info, portfolio_path, new_shares
    )

    return analysis


# ─────────────────────────────────────────────
# FORMAT FOR LLM
# ─────────────────────────────────────────────

def format_bogle_for_llm(analysis: BogleAnalysis) -> str:
    lw = analysis.last_war
    rv = analysis.reversion
    da = analysis.diversification

    def p(val, d=1):
        return f"{val:.{d}%}" if val is not None else "N/A"

    def n(val, d=2):
        return f"{val:.{d}f}" if val is not None else "N/A"

    def dollar(val):
        if val is None: return "N/A"
        if abs(val) >= 1e6: return f"${val/1e6:.1f}M"
        if abs(val) >= 1e3: return f"${val/1e3:.0f}K"
        return f"${val:.0f}"

    # Sector table — value
    value_rows = ""
    for s, pct in sorted(da.sector_by_value.items(), key=lambda x: -x[1]):
        after_pct = da.after_sector_by_value.get(s, pct)
        arrow = "▲" if after_pct > pct + 0.005 else ("▼" if after_pct < pct - 0.005 else "—")
        flag = " 🔴" if pct >= SECTOR_CRITICAL else (" ⚠️" if pct >= SECTOR_WARN_PCT else "")
        value_rows += f"  {s:<30} {pct:>7.1%}   →   {after_pct:>7.1%}  {arrow}{flag}\n"
    # New sector if not in portfolio
    if da.new_ticker_sector not in da.sector_by_value and da.new_position_value > 0:
        after_pct = da.after_sector_by_value.get(da.new_ticker_sector, 0)
        value_rows += f"  {da.new_ticker_sector:<30} {'0.0%':>7}   →   {after_pct:>7.1%}  ▲ NEW\n"

    # Sector table — income
    income_rows = ""
    for s, pct in sorted(da.sector_by_income.items(), key=lambda x: -x[1]):
        after_pct = da.after_sector_by_income.get(s, pct)
        arrow = "▲" if after_pct > pct + 0.005 else ("▼" if after_pct < pct - 0.005 else "—")
        flag = " ⚠️" if pct >= INCOME_WARN_PCT else ""
        income_rows += f"  {s:<30} {pct:>7.1%}   →   {after_pct:>7.1%}  {arrow}{flag}\n"
    if da.new_ticker_sector not in da.sector_by_income and da.new_annual_income > 0:
        after_pct = da.after_sector_by_income.get(da.new_ticker_sector, 0)
        income_rows += f"  {da.new_ticker_sector:<30} {'0.0%':>7}   →   {after_pct:>7.1%}  ▲ NEW\n"

    # Pre-compute warning strings — Python 3.9 can't use backslash inside f-string {}
    NL = "\n"
    value_warn_str = ("VALUE WARNINGS:\n" + NL.join(da.value_warnings)) if da.value_warnings else "✅ No value concentration warnings."
    income_warn_str = ("INCOME WARNINGS:\n" + NL.join(da.income_warnings)) if da.income_warnings else "✅ No income concentration warnings."

    block = f"""
================================================================================
BOGLE ANALYSIS FACT SHEET — {analysis.ticker} ({analysis.company_name})
Sector: {analysis.sector}
================================================================================

YOUR ROLE: You are John Bogle. Interpret the data below through your three
core principles. Reference the specific numbers. Be direct and brief.
Do NOT invent data. Do not repeat what the table already says — synthesize it.

────────────────────────────────────────────────────────────────────────────────
PILLAR 1 — PAST PERFORMANCE RELIABILITY
(Is this return repeatable, or was it a one-time regime?)
────────────────────────────────────────────────────────────────────────────────
5yr Annualized Return:   {p(lw.annualized_return_pct)}
5yr Annualized Volatility:{p(lw.annualized_volatility)}
Sharpe Ratio:            {n(lw.sharpe_ratio)}    (> 1.0 = good risk-adjusted return)
Max Drawdown:            {p(lw.max_drawdown)}
Beta vs S&P 500:         {n(lw.beta_to_spy)}    (1.0 = market, > 1.5 = highly cyclical)
Return vs S&P 500:       {p(lw.return_vs_spy)}  per year (alpha/underperformance)

Performance Regime:      {lw.performance_regime}
Past Perf. Risk:           {lw.last_war_risk}
Assessment:              {lw.last_war_warning}

────────────────────────────────────────────────────────────────────────────────
PILLAR 2 — REVERSION TO THE MEAN
(Buy timing: BEST / GOOD / MEDIUM / BAD)
────────────────────────────────────────────────────────────────────────────────
Current Price:           ${n(rv.price_current)}
50-day MA:               ${n(rv.ma_50d)}   ({p(rv.pct_above_ma50)} above/below)
200-day MA:              ${n(rv.ma_200d)}  ({p(rv.pct_above_ma200)} above/below)
Z-Score vs 1yr mean:     {n(rv.z_score_1yr)}   (0 = at mean, +2 = 2 std devs above)
Z-Score vs 3yr mean:     {n(rv.z_score_3yr)}
RSI (14-day):            {n(rv.rsi_14, 0)}    (< 35 = oversold, > 70 = overbought)
P/E Current:             {n(rv.pe_current, 1)}x

Timing Score:            {rv.timing_score}/10
BUY TIMING SIGNAL:       {rv.timing_signal}
Reasoning:               {rv.timing_reasoning}

────────────────────────────────────────────────────────────────────────────────
PILLAR 3 — PORTFOLIO DIVERSIFICATION
Assuming adding {da.new_shares_assumed} shares of {da.new_ticker} (${da.new_position_value:,.0f} value, ${da.new_annual_income:.0f}/yr income)
────────────────────────────────────────────────────────────────────────────────

CURRENT PORTFOLIO:  ${da.total_value:,.0f} total value  |  ${da.total_annual_income:,.0f}/yr income
Diversification Grade: {da.diversification_grade}  (HHI: {n(da.herfindahl_index, 3)})

SECTOR ALLOCATION BY PORTFOLIO VALUE        Before  →   After
{value_rows}
SECTOR ALLOCATION BY DIVIDEND INCOME        Before  →   After
{income_rows}
{value_warn_str}
{income_warn_str}

DIVERSIFICATION IMPACT:  {da.diversification_impact}
Impact Reasoning:        {da.impact_reasoning}

================================================================================
INSTRUCTION — John Bogle, answer ALL of these specifically:

1. PAST PERFORMANCE RELIABILITY: Given the Sharpe ratio, beta, and max drawdown shown,
   is this return repeatable or was it a regime fluke? Be specific — don't just say "cyclical."

2. WEIGHTED PORTFOLIO COST CHECK: Johnathan's current holdings include individual stocks
   (O, MSFT, VICI, PFE, GIS, UPS). Estimate what his blended portfolio cost drag likely is
   vs a 3-fund index portfolio (VTI/VXUS/BND). How much is that friction costing him per year
   on a $50,000 portfolio over 30 years at 7% growth?

3. INDEX ALTERNATIVE: What specific index fund or ETF would give Johnathan similar exposure
   to this stock with lower cost and better diversification? Name the ticker.

4. DIVERSIFICATION VERDICT: Does adding this ACTUALLY reduce his HHI concentration score,
   or is he kidding himself? State the before/after HHI numbers.

5. BUY TIMING: Is this a good entry point by mean reversion? Give a specific price target
   where the reversion math would make it more attractive.
================================================================================
"""
    return block


def format_bogle_display(analysis: BogleAnalysis) -> str:
    """Compact display for the chat window."""
    lw = analysis.last_war
    rv = analysis.reversion
    da = analysis.diversification

    def p(val): return f"{val:.1%}" if val is not None else "N/A"
    def n(val, d=2): return f"{val:.{d}f}" if val is not None else "N/A"

    timing_bar = {"BEST": "████████████ BEST", "GOOD": "█████████░░░ GOOD",
                  "MEDIUM": "██████░░░░░░ MEDIUM", "BAD": "███░░░░░░░░░ BAD"}.get(rv.timing_signal, rv.timing_signal)

    lines = [
        f"",
        f"  ── PAST PERFORMANCE RELIABILITY ──────────────────────────",
        f"  5yr Return:    {p(lw.annualized_return_pct)}   Vol: {p(lw.annualized_volatility)}   Sharpe: {n(lw.sharpe_ratio)}",
        f"  Beta vs SPY:   {n(lw.beta_to_spy)}   Max Drawdown: {p(lw.max_drawdown)}",
        f"  Regime:        {lw.performance_regime}   Risk: {lw.last_war_risk}",
        f"",
        f"  ── REVERSION TO MEAN ──────────────────────────────────────",
        f"  Z-Score 3yr:   {n(rv.z_score_3yr)}   RSI: {n(rv.rsi_14, 0)}",
        f"  vs 200d MA:    {p(rv.pct_above_ma200)}   vs 50d MA: {p(rv.pct_above_ma50)}",
        f"  Timing:        {timing_bar}  ({rv.timing_score}/10)",
        f"",
        f"  ── PORTFOLIO IMPACT ────────────────────────────────────────",
        f"  New Sector:    {da.new_ticker_sector}",
        f"  Impact:        {da.diversification_impact}",
        f"  Grade Before:  {da.diversification_grade}",
    ]
    return "\n".join(lines)


# ─────────────────────────────────────────────
# QUICK TEST
# ─────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    ticker = sys.argv[1] if len(sys.argv) > 1 else "PFE"
    print(f"\nRunning Bogle analysis on {ticker}...\n")
    result = run_bogle_analysis(ticker)
    print(format_bogle_display(result))
    print()
    print(format_bogle_for_llm(result))
