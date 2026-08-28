"""
dalio_analyzer.py
==================
Ray Dalio's framework — four concrete filters applied before the LLM sees anything.

A. Debt-to-Income Filter    (Big Debt Cycle)
B. Productivity vs Cost     (Efficiency / Decline Phase)
C. Bubble Indicator         (3-point automated bubble check)
D. Correlation Filter       (Holy Grail — uncorrelated assets)

Requires: pip install yfinance pandas numpy
"""

import yfinance as yf
import pandas as pd
import numpy as np
from dataclasses import dataclass, field
from typing import Optional
from datetime import datetime
import os

# ─────────────────────────────────────────────
# DALIO THRESHOLDS
# ─────────────────────────────────────────────
DEBT_EBITDA_MAX         = 3.0    # Total Debt / EBITDA < 3.0
PROD_SPREAD_MIN         = 0.0    # Revenue growth - OpEx growth must be positive
PE_FROTHY_MULTIPLIER    = 1.5    # Current P/E > 1.5x 5yr avg = frothy
PEG_BUBBLE_MAX          = 2.5    # PEG > 2.5 = unsustainable growth priced in
FCF_QUALITY_MIN         = 0.80   # FCF / Net Income < 0.8 = accounting concern
CORRELATION_MAX         = 0.30   # correlation > 0.3 = not truly uncorrelated
HOLY_GRAIL_TARGET       = 15     # Dalio's 15+ uncorrelated assets goal

# Benchmark and portfolio tickers for correlation
BENCHMARK_TICKERS       = ["SPY", "QQQ", "TLT", "GLD"]   # broad market proxies
CORRELATION_LOOKBACK    = "3y"   # 3 years of daily returns for correlation


# ─────────────────────────────────────────────
# DATA STRUCTURES
# ─────────────────────────────────────────────

@dataclass
class FilterResult:
    name:           str = ""
    passed:         bool = False
    metric_name:    str = ""
    actual_value:   Optional[float] = None
    threshold:      str = ""
    display_value:  str = ""
    cycle_context:  str = ""    # Dalio's macro framing
    note:           str = ""


@dataclass
class DebtCycleFilter:
    total_debt:         Optional[float] = None
    ebitda:             Optional[float] = None
    debt_to_ebitda:     Optional[float] = None
    result:             FilterResult = field(default_factory=FilterResult)


@dataclass
class ProductivityFilter:
    revenue_growth:     Optional[float] = None
    opex_growth:        Optional[float] = None
    spread:             Optional[float] = None   # revenue growth - opex growth
    result:             FilterResult = field(default_factory=FilterResult)


@dataclass
class BubbleFilter:
    # Check 1: Valuation extremes
    pe_current:         Optional[float] = None
    pe_5yr_avg:         Optional[float] = None
    pe_ratio_vs_avg:    Optional[float] = None   # current / 5yr avg

    # Check 2: Unsustainable growth
    peg_ratio:          Optional[float] = None

    # Check 3: FCF quality
    free_cash_flow:     Optional[float] = None
    net_income:         Optional[float] = None
    fcf_to_ni:          Optional[float] = None

    checks_passed:      int = 0
    checks_total:       int = 3
    result:             FilterResult = field(default_factory=FilterResult)
    sub_results:        list = field(default_factory=list)   # list of FilterResult


@dataclass
class CorrelationResult:
    ticker:             str = ""
    corr_to_spy:        Optional[float] = None
    corr_to_qqq:        Optional[float] = None
    corr_to_tlt:        Optional[float] = None
    corr_to_gld:        Optional[float] = None

    # Correlation to existing portfolio holdings
    portfolio_correlations: dict = field(default_factory=dict)   # ticker -> correlation

    max_portfolio_corr: Optional[float] = None   # worst (highest) correlation to any holding
    avg_portfolio_corr: Optional[float] = None
    uncorrelated_count: int = 0     # number of holdings with corr < 0.3

    # Holy Grail assessment
    adds_diversification:   bool = False
    holy_grail_verdict:     str = ""
    result:                 FilterResult = field(default_factory=FilterResult)


@dataclass
class DalioAnalysis:
    ticker:             str = ""
    company_name:       str = ""
    sector:             str = ""
    analysis_date:      str = ""

    debt_cycle:         DebtCycleFilter = field(default_factory=DebtCycleFilter)
    productivity:       ProductivityFilter = field(default_factory=ProductivityFilter)
    bubble:             BubbleFilter = field(default_factory=BubbleFilter)
    correlation:        CorrelationResult = field(default_factory=CorrelationResult)

    filters_passed:     int = 0
    filters_total:      int = 4
    overall_signal:     str = "UNKNOWN"   # PASS / PARTIAL / FAIL
    errors:             list = field(default_factory=list)


# ─────────────────────────────────────────────
# FILTER A: DEBT-TO-INCOME (BIG DEBT CYCLE)
# ─────────────────────────────────────────────

def analyze_debt_cycle(ticker: str, info: dict) -> DebtCycleFilter:
    dc = DebtCycleFilter()
    r = FilterResult(
        name="A. Debt-to-Income (Big Debt Cycle)",
        metric_name="Total Debt / EBITDA",
        threshold=f"< {DEBT_EBITDA_MAX}x",
        cycle_context=(
            "In a late-cycle high-rate environment, companies with Debt/EBITDA > 3x "
            "face refinancing risk and earnings compression. Dalio's template: "
            "debt cannot grow faster than income."
        )
    )

    try:
        total_debt = info.get("totalDebt")
        ebitda = info.get("ebitda")

        dc.total_debt = float(total_debt) if total_debt else None
        dc.ebitda = float(ebitda) if ebitda else None

        if dc.total_debt is not None and dc.ebitda and dc.ebitda > 0:
            dc.debt_to_ebitda = dc.total_debt / dc.ebitda
            r.actual_value = dc.debt_to_ebitda
            r.display_value = f"{dc.debt_to_ebitda:.2f}x"
            r.passed = dc.debt_to_ebitda < DEBT_EBITDA_MAX

            if r.passed:
                r.note = f"Debt/EBITDA of {dc.debt_to_ebitda:.2f}x — manageable leverage. Survives rate pressure."
            else:
                r.note = (
                    f"Debt/EBITDA of {dc.debt_to_ebitda:.2f}x exceeds 3.0x threshold. "
                    f"Dalio: first to fall in a debt deleveraging cycle."
                )
        elif dc.total_debt == 0 or dc.total_debt is None:
            r.display_value = "No debt"
            r.passed = True
            r.note = "No significant debt — ideal from a debt cycle perspective."
            dc.debt_to_ebitda = 0.0
        else:
            r.display_value = "N/A"
            r.passed = False
            r.note = "EBITDA unavailable — cannot assess debt serviceability."

    except Exception as e:
        r.display_value = "ERROR"
        r.note = str(e)

    dc.result = r
    return dc


# ─────────────────────────────────────────────
# FILTER B: PRODUCTIVITY vs COST
# ─────────────────────────────────────────────

def analyze_productivity(ticker: str, info: dict) -> ProductivityFilter:
    pf = ProductivityFilter()
    r = FilterResult(
        name="B. Productivity vs Cost",
        metric_name="Revenue Growth % − Operating Expense Growth %",
        threshold="> 0% (positive spread)",
        cycle_context=(
            "If costs grow faster than revenue, the company is in decline phase. "
            "Dalio calls this 'getting fat and happy' — a peak-cycle warning sign."
        )
    )

    try:
        t = yf.Ticker(ticker)
        income = t.income_stmt

        if income is not None and not income.empty and len(income.columns) >= 2:
            # Get 2 most recent annual periods
            rev_key = None
            for k in ["Total Revenue", "Revenue"]:
                if k in income.index:
                    rev_key = k
                    break

            opex_key = None
            for k in ["Operating Expense", "Total Operating Expenses", "Operating Expenses"]:
                if k in income.index:
                    opex_key = k
                    break

            if rev_key and opex_key:
                rev = income.loc[rev_key]
                opex = income.loc[opex_key]

                rev_recent  = float(rev.iloc[0])
                rev_prior   = float(rev.iloc[1])
                opex_recent = float(opex.iloc[0])
                opex_prior  = float(opex.iloc[1])

                if rev_prior != 0:
                    pf.revenue_growth = (rev_recent - rev_prior) / abs(rev_prior)
                if opex_prior != 0:
                    pf.opex_growth = (opex_recent - opex_prior) / abs(opex_prior)

                if pf.revenue_growth is not None and pf.opex_growth is not None:
                    pf.spread = pf.revenue_growth - pf.opex_growth
                    r.actual_value = pf.spread
                    r.display_value = f"{pf.spread:+.1%} spread (rev {pf.revenue_growth:+.1%} − opex {pf.opex_growth:+.1%})"
                    r.passed = pf.spread > PROD_SPREAD_MIN

                    if r.passed:
                        r.note = (
                            f"Revenue growing {pf.spread:.1%} faster than costs. "
                            f"Efficiency improving — not in decline phase."
                        )
                    else:
                        r.note = (
                            f"Costs growing {abs(pf.spread):.1%} faster than revenue. "
                            f"Dalio: margin compression ahead, company is in decline phase."
                        )
                else:
                    r.display_value = "N/A — growth rates unavailable"
                    r.passed = False
                    r.note = "Cannot compute productivity spread from available data."
            else:
                # Fallback: use yfinance info fields
                rev_growth = info.get("revenueGrowth")
                op_margins_change = info.get("operatingMargins")
                if rev_growth is not None:
                    pf.revenue_growth = float(rev_growth)
                    r.display_value = f"Rev growth: {pf.revenue_growth:+.1%} (opex breakdown unavailable)"
                    r.passed = pf.revenue_growth > 0
                    r.note = "Only revenue growth available — positive revenue growth used as proxy."
                    r.actual_value = pf.revenue_growth
                else:
                    r.display_value = "N/A"
                    r.passed = False
                    r.note = "Income statement data unavailable for productivity calculation."
        else:
            r.display_value = "N/A"
            r.passed = False
            r.note = "Income statement not available."

    except Exception as e:
        r.display_value = "ERROR"
        r.note = str(e)
        r.passed = False

    pf.result = r
    return pf


# ─────────────────────────────────────────────
# FILTER C: BUBBLE INDICATOR (3-point check)
# ─────────────────────────────────────────────

def get_5yr_avg_pe(ticker: str) -> Optional[float]:
    """
    Estimate 5yr average P/E from historical price and EPS data.
    Uses annual EPS from income statement and year-end prices.
    """
    try:
        t = yf.Ticker(ticker)
        hist = t.history(period="5y", interval="3mo")
        income = t.income_stmt

        if hist.empty or income is None or income.empty:
            return None

        eps_key = None
        for k in ["Basic EPS", "Diluted EPS", "EPS"]:
            if k in income.index:
                eps_key = k
                break

        if not eps_key:
            return None

        eps_series = income.loc[eps_key].dropna()
        if len(eps_series) < 2:
            return None

        pes = []
        for date, eps in eps_series.items():
            if eps and float(eps) > 0:
                # Find closest price to this date
                try:
                    date_naive = date.replace(tzinfo=None) if hasattr(date, 'tzinfo') else date
                    hist_naive = hist.copy()
                    hist_naive.index = hist_naive.index.tz_localize(None) if hist_naive.index.tz else hist_naive.index
                    idx = hist_naive.index.get_indexer([date_naive], method='nearest')[0]
                    if idx >= 0:
                        price = float(hist_naive['Close'].iloc[idx])
                        pe = price / float(eps)
                        if 0 < pe < 200:   # sanity check
                            pes.append(pe)
                except Exception:
                    continue

        return float(np.mean(pes)) if pes else None

    except Exception:
        return None


def analyze_bubble(ticker: str, info: dict) -> BubbleFilter:
    bf = BubbleFilter()
    sub_results = []

    # ── Check 1: Valuation Extremes ──────────────────────────────────────────
    r1 = FilterResult(
        name="C1. Valuation Extremes (P/E vs 5yr avg)",
        metric_name="Current P/E / 5yr Avg P/E",
        threshold=f"< {PE_FROTHY_MULTIPLIER}x ratio",
        cycle_context="If current P/E > 1.5x its 5yr average, the market is pricing in perfection — Dalio bubble signal."
    )

    bf.pe_current = info.get("trailingPE")
    bf.pe_5yr_avg = get_5yr_avg_pe(ticker)

    if bf.pe_current and bf.pe_5yr_avg and bf.pe_5yr_avg > 0:
        bf.pe_ratio_vs_avg = bf.pe_current / bf.pe_5yr_avg
        r1.actual_value = bf.pe_ratio_vs_avg
        r1.display_value = f"{bf.pe_ratio_vs_avg:.2f}x (current P/E {bf.pe_current:.1f}x vs 5yr avg {bf.pe_5yr_avg:.1f}x)"
        r1.passed = bf.pe_ratio_vs_avg < PE_FROTHY_MULTIPLIER
        r1.note = ("Valuation in line with history — not frothy." if r1.passed else
                   f"Current P/E is {bf.pe_ratio_vs_avg:.1f}x its 5yr avg — Dalio frothy signal. Market pricing in perfection.")
    elif bf.pe_current and not bf.pe_5yr_avg:
        r1.display_value = f"P/E {bf.pe_current:.1f}x (5yr avg unavailable)"
        r1.passed = bf.pe_current < 25   # fallback: raw P/E < 25 = not extreme
        r1.note = "5yr avg P/E unavailable — using absolute P/E < 25 as proxy."
    else:
        r1.display_value = "N/A"
        r1.passed = False
        r1.note = "P/E data unavailable."
    sub_results.append(r1)

    # ── Check 2: Unsustainable Growth (PEG) ──────────────────────────────────
    r2 = FilterResult(
        name="C2. Unsustainable Growth (PEG Ratio)",
        metric_name="PEG Ratio",
        threshold=f"< {PEG_BUBBLE_MAX}",
        cycle_context="PEG > 2.5 means the market is paying for growth that cannot realistically materialize — Dalio: 'discounting unsustainable perfection'."
    )

    bf.peg_ratio = info.get("pegRatio")
    if bf.peg_ratio:
        peg = float(bf.peg_ratio)
        r2.actual_value = peg
        r2.display_value = f"{peg:.2f}"
        r2.passed = peg < PEG_BUBBLE_MAX
        r2.note = ("PEG within reasonable range — growth is credibly priced." if r2.passed else
                   f"PEG of {peg:.2f} signals unsustainable growth expectations — bubble risk.")
    else:
        r2.display_value = "N/A"
        r2.passed = True   # benefit of the doubt if no data
        r2.note = "PEG unavailable (often N/A for dividend stocks / value plays) — defaulting to pass."
    sub_results.append(r2)

    # ── Check 3: FCF Quality ─────────────────────────────────────────────────
    r3 = FilterResult(
        name="C3. Cash Flow Reality (FCF / Net Income)",
        metric_name="Free Cash Flow / Net Income",
        threshold=f">= {FCF_QUALITY_MIN:.0%}",
        cycle_context="If FCF < 80% of net income, earnings are accounting entries not real cash. Dalio: 'creative accounting' spikes at cycle peaks."
    )

    bf.free_cash_flow = info.get("freeCashflow")
    bf.net_income = info.get("netIncomeToCommon") or info.get("netIncome")

    if bf.free_cash_flow and bf.net_income and float(bf.net_income) > 0:
        bf.fcf_to_ni = float(bf.free_cash_flow) / float(bf.net_income)
        r3.actual_value = bf.fcf_to_ni
        r3.display_value = f"{bf.fcf_to_ni:.1%}"
        r3.passed = bf.fcf_to_ni >= FCF_QUALITY_MIN
        r3.note = ("FCF closely tracks earnings — accounting appears clean." if r3.passed else
                   f"FCF is only {bf.fcf_to_ni:.0%} of net income — Dalio creative accounting flag.")
    elif bf.free_cash_flow and bf.net_income and float(bf.net_income) < 0:
        r3.display_value = "Net loss (negative NI)"
        r3.passed = False
        r3.note = "Company is unprofitable — FCF/NI ratio not meaningful."
    else:
        r3.display_value = "N/A"
        r3.passed = False
        r3.note = "FCF or net income data unavailable."
    sub_results.append(r3)

    # ── Aggregate bubble score ────────────────────────────────────────────────
    bf.sub_results = sub_results
    bf.checks_passed = sum(1 for r in sub_results if r.passed)

    overall = FilterResult(
        name="C. Bubble Indicator (3-point check)",
        metric_name="Bubble checks passed",
        threshold="All 3 passing",
    )
    overall.actual_value = bf.checks_passed
    overall.display_value = f"{bf.checks_passed}/{bf.checks_total} checks passing"
    overall.passed = bf.checks_passed == bf.checks_total

    if bf.checks_passed == 3:
        overall.note = "No bubble signals detected across valuation, growth pricing, and accounting quality."
    elif bf.checks_passed == 2:
        overall.note = "One bubble signal present — elevated but not extreme risk."
    elif bf.checks_passed == 1:
        overall.note = "Two bubble signals — Dalio would be cautious. High cycle risk."
    else:
        overall.note = "All three bubble signals triggered — Dalio would avoid. Classic late-cycle setup."

    bf.result = overall
    return bf


# ─────────────────────────────────────────────
# FILTER D: CORRELATION (HOLY GRAIL)
# ─────────────────────────────────────────────

def load_portfolio_tickers(portfolio_path: str = "portfolio.xlsx") -> list:
    """Load current holdings tickers from Excel."""
    tickers = []
    if not os.path.exists(portfolio_path):
        return tickers
    try:
        import openpyxl
        wb = openpyxl.load_workbook(portfolio_path, data_only=True)
        if "Holdings" not in wb.sheetnames:
            return tickers
        ws = wb["Holdings"]
        for row in ws.iter_rows(min_row=5, values_only=True):
            t = row[0]
            if t and str(t).strip() not in ("", "TOTALS", "CASH"):
                tickers.append(str(t).strip().upper())
    except Exception:
        pass
    return tickers


def analyze_correlation(ticker: str, portfolio_path: str = "portfolio.xlsx") -> CorrelationResult:
    cr = CorrelationResult(ticker=ticker.upper())

    r = FilterResult(
        name="D. Correlation Filter (Holy Grail)",
        metric_name="Max correlation to existing holdings",
        threshold=f"< {CORRELATION_MAX} (uncorrelated)",
        cycle_context=(
            "Dalio's Holy Grail: 15+ truly uncorrelated return streams. "
            "Correlation > 0.3 to existing holdings means 'fake diversification' — "
            "you're just adding more of what you already have."
        )
    )

    try:
        # Get portfolio tickers + benchmarks
        portfolio_tickers = load_portfolio_tickers(portfolio_path)
        all_comparison = list(set(portfolio_tickers + BENCHMARK_TICKERS))

        if not all_comparison:
            r.display_value = "No portfolio data"
            r.passed = True
            r.note = "No existing portfolio to correlate against — cannot run Holy Grail check."
            cr.result = r
            return cr

        # Download price history for new ticker + all comparison tickers
        all_tickers = [ticker.upper()] + [t for t in all_comparison if t != ticker.upper()]

        data = yf.download(all_tickers, period=CORRELATION_LOOKBACK,
                           progress=False, auto_adjust=True)

        if "Close" in data:
            prices = data["Close"]
        else:
            prices = data

        if prices.empty or ticker.upper() not in prices.columns:
            r.display_value = "Download failed"
            r.passed = False
            r.note = "Could not download price history for correlation analysis."
            cr.result = r
            return cr

        # Calculate daily returns correlation matrix
        returns = prices.pct_change().dropna()
        corr_matrix = returns.corr()

        if ticker.upper() not in corr_matrix.index:
            r.display_value = "N/A"
            r.passed = False
            cr.result = r
            return cr

        ticker_corrs = corr_matrix[ticker.upper()]

        # Benchmark correlations
        for bm in BENCHMARK_TICKERS:
            if bm in ticker_corrs.index:
                val = float(ticker_corrs[bm])
                setattr(cr, f"corr_to_{bm.lower()}", val)

        # Portfolio holding correlations
        for pt in portfolio_tickers:
            if pt in ticker_corrs.index and pt != ticker.upper():
                cr.portfolio_correlations[pt] = float(ticker_corrs[pt])

        if cr.portfolio_correlations:
            cr.max_portfolio_corr = max(cr.portfolio_correlations.values())
            cr.avg_portfolio_corr = float(np.mean(list(cr.portfolio_correlations.values())))
            cr.uncorrelated_count = sum(1 for v in cr.portfolio_correlations.values()
                                        if v < CORRELATION_MAX)

            r.actual_value = cr.max_portfolio_corr
            r.passed = cr.max_portfolio_corr < CORRELATION_MAX

            # Build correlation breakdown string
            corr_lines = []
            for pt, corr in sorted(cr.portfolio_correlations.items(), key=lambda x: -x[1]):
                flag = " ⚠️ HIGH" if corr >= CORRELATION_MAX else " ✅"
                corr_lines.append(f"{pt}: {corr:.2f}{flag}")
            r.display_value = f"Max corr to holdings: {cr.max_portfolio_corr:.2f}"

            if r.passed:
                cr.adds_diversification = True
                cr.holy_grail_verdict = (
                    f"LOW correlation to all existing holdings (max {cr.max_portfolio_corr:.2f}). "
                    f"Dalio: this adds a genuine return stream — true diversification."
                )
                r.note = cr.holy_grail_verdict
            else:
                # Find which holdings it's highly correlated to
                high_corr = [f"{t} ({v:.2f})" for t, v in cr.portfolio_correlations.items()
                             if v >= CORRELATION_MAX]
                cr.adds_diversification = False
                cr.holy_grail_verdict = (
                    f"HIGH correlation to: {', '.join(high_corr)}. "
                    f"Dalio: 'fake diversification' — you already own this risk."
                )
                r.note = cr.holy_grail_verdict
        else:
            r.display_value = "No portfolio overlap"
            r.passed = True
            r.note = "No matching portfolio tickers for correlation — checking benchmarks only."

        # Benchmark note
        bm_notes = []
        for bm in BENCHMARK_TICKERS:
            val = getattr(cr, f"corr_to_{bm.lower()}", None)
            if val is not None:
                bm_notes.append(f"{bm}: {val:.2f}")
        if bm_notes:
            r.note += f" | Benchmark correlations: {', '.join(bm_notes)}"

    except Exception as e:
        r.display_value = "ERROR"
        r.note = str(e)
        r.passed = False

    cr.result = r
    return cr


# ─────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────

def run_dalio_analysis(ticker: str, portfolio_path: str = "portfolio.xlsx") -> DalioAnalysis:
    analysis = DalioAnalysis(
        ticker=ticker.upper(),
        analysis_date=datetime.now().strftime("%Y-%m-%d")
    )

    try:
        t = yf.Ticker(ticker)
        info = t.info
        analysis.company_name = info.get("longName", ticker)
        analysis.sector = info.get("sector", "Unknown")
    except Exception as e:
        analysis.errors.append(f"Info fetch failed: {e}")
        info = {}

    # Run all four filters
    analysis.debt_cycle   = analyze_debt_cycle(ticker, info)
    analysis.productivity = analyze_productivity(ticker, info)
    analysis.bubble       = analyze_bubble(ticker, info)
    analysis.correlation  = analyze_correlation(ticker, portfolio_path)

    # Overall signal
    results = [
        analysis.debt_cycle.result,
        analysis.productivity.result,
        analysis.bubble.result,
        analysis.correlation.result,
    ]
    analysis.filters_passed = sum(1 for r in results if r.passed)

    if analysis.filters_passed == 4:
        analysis.overall_signal = "PASS — All Dalio filters satisfied"
    elif analysis.filters_passed >= 3:
        analysis.overall_signal = "PARTIAL — Minor concerns, review flagged filters"
    elif analysis.filters_passed >= 2:
        analysis.overall_signal = "CAUTION — Multiple Dalio filters failing"
    else:
        analysis.overall_signal = "FAIL — Dalio would reject this position"

    return analysis


# ─────────────────────────────────────────────
# FORMAT FOR LLM
# ─────────────────────────────────────────────

def format_dalio_for_llm(analysis: DalioAnalysis) -> str:
    dc = analysis.debt_cycle
    pf = analysis.productivity
    bf = analysis.bubble
    cr = analysis.correlation

    def p(val, d=1):
        return f"{val:+.{d}%}" if val is not None else "N/A"

    def n(val, d=2):
        return f"{val:.{d}f}" if val is not None else "N/A"

    # Sub-results for bubble
    bubble_detail = ""
    for sr in bf.sub_results:
        sym = "✅" if sr.passed else "❌"
        bubble_detail += f"  {sym} {sr.name}: {sr.display_value}\n"
        bubble_detail += f"       {sr.note}\n"

    # Correlation table
    corr_table = ""
    if cr.portfolio_correlations:
        for pt, corr in sorted(cr.portfolio_correlations.items(), key=lambda x: -x[1]):
            flag = "⚠️  HIGH" if corr >= CORRELATION_MAX else "✅ OK  "
            corr_table += f"  {flag}  {pt:<8} corr = {corr:.3f}\n"

    block = f"""
================================================================================
DALIO ANALYSIS FACT SHEET — {analysis.ticker} ({analysis.company_name})
Sector: {analysis.sector} | Date: {analysis.analysis_date}
Overall: {analysis.filters_passed}/4 filters passing — {analysis.overall_signal}
================================================================================

YOUR ROLE: You are Ray Dalio. Interpret these four filter results through your
framework of debt cycles, productivity, bubble detection, and the Holy Grail.
Reference specific numbers. Do NOT invent data. Max 200 words. No pleasantries.
Open with your overall verdict on whether this fits Johnathan's portfolio.

────────────────────────────────────────────────────────────────────────────────
FILTER A — DEBT-TO-INCOME (Big Debt Cycle)
────────────────────────────────────────────────────────────────────────────────
{"✅ PASS" if dc.result.passed else "❌ FAIL"}  Total Debt / EBITDA: {dc.result.display_value}
Threshold: < 3.0x
Total Debt: {"${:,.0f}".format(dc.total_debt) if dc.total_debt else "N/A"}  |  EBITDA: {"${:,.0f}".format(dc.ebitda) if dc.ebitda else "N/A"}
Dalio Context: {dc.result.cycle_context}
Assessment: {dc.result.note}

────────────────────────────────────────────────────────────────────────────────
FILTER B — PRODUCTIVITY vs COST (Decline Phase Check)
────────────────────────────────────────────────────────────────────────────────
{"✅ PASS" if pf.result.passed else "❌ FAIL"}  Spread: {pf.result.display_value}
Revenue Growth: {p(pf.revenue_growth)}  |  OpEx Growth: {p(pf.opex_growth)}
Threshold: Revenue must grow faster than costs (positive spread)
Dalio Context: {pf.result.cycle_context}
Assessment: {pf.result.note}

────────────────────────────────────────────────────────────────────────────────
FILTER C — BUBBLE INDICATOR ({bf.checks_passed}/{bf.checks_total} sub-checks passing)
────────────────────────────────────────────────────────────────────────────────
{bubble_detail}
Overall Bubble Assessment: {bf.result.note}

────────────────────────────────────────────────────────────────────────────────
FILTER D — CORRELATION / HOLY GRAIL
────────────────────────────────────────────────────────────────────────────────
{"✅ PASS" if cr.result.passed else "❌ FAIL"}  {cr.result.display_value}
Threshold: Max correlation to any holding < {CORRELATION_MAX}
Lookback: {CORRELATION_LOOKBACK} daily returns

Correlation to existing holdings:
{corr_table if corr_table else "  [No portfolio data available]"}
Max portfolio correlation: {n(cr.max_portfolio_corr)}
Avg portfolio correlation: {n(cr.avg_portfolio_corr)}

Holy Grail Verdict: {cr.holy_grail_verdict}
Benchmark correlations: SPY {n(cr.corr_to_spy)} | QQQ {n(cr.corr_to_qqq)} | TLT {n(cr.corr_to_tlt)} | GLD {n(cr.corr_to_gld)}

================================================================================
INSTRUCTION — Ray Dalio, answer ALL of these specifically:

1. DEBT CYCLE POSITION: Based on Debt/EBITDA and the current macro environment
   (Buffett Indicator at extreme highs, elevated rates), where does this company sit
   in the debt cycle template? Early/mid/late/deleveraging — and what happens to it
   in each scenario?

2. PRODUCTIVITY TREND JUDGMENT: Is the revenue-vs-cost spread widening or narrowing?
   If narrowing, how many more years at this rate before margins compress to zero?
   Use the actual numbers to estimate this.

3. BUBBLE RISK SPECIFICS: Of the 3 bubble checks, which failure concerns you most and why?
   Quantify the risk — if the P/E reverts to its 5yr average, what price does that imply?

4. CORRELATION VERDICT WITH SPECIFICS: Name the existing holding this stock correlates
   most highly with. Explain in one sentence what that means for Johnathan's actual risk —
   is he doubling up on a risk factor he can't see?

5. HOLY GRAIL ALTERNATIVE: If this stock fails the correlation filter, name one specific
   asset class or ticker that WOULD add a genuinely uncorrelated return stream to his
   current portfolio. Be specific — not just "bonds," give a ticker.

End with: PASS / PARTIAL / FAIL and one sentence on what would need to change to flip the verdict.
================================================================================
"""
    return block


def format_dalio_display(analysis: DalioAnalysis) -> str:
    """Compact display for chat window."""
    dc = analysis.debt_cycle
    pf = analysis.productivity
    bf = analysis.bubble
    cr = analysis.correlation

    def n(val, d=2):
        return f"{val:.{d}f}" if val is not None else "N/A"

    lines = [
        "",
        f"  ── FILTER A: DEBT CYCLE ────────────────────────────────",
        f"  {'✅' if dc.result.passed else '❌'} Debt/EBITDA: {dc.result.display_value:<20} (threshold: < 3.0x)",
        "",
        f"  ── FILTER B: PRODUCTIVITY ──────────────────────────────",
        f"  {'✅' if pf.result.passed else '❌'} {pf.result.display_value}",
        "",
        f"  ── FILTER C: BUBBLE ({bf.checks_passed}/{bf.checks_total} checks) ─────────────────────────",
    ]
    for sr in bf.sub_results:
        lines.append(f"  {'✅' if sr.passed else '❌'} {sr.name:<42} {sr.display_value}")
    lines += [
        "",
        f"  ── FILTER D: HOLY GRAIL CORRELATION ────────────────────",
        f"  {'✅' if cr.result.passed else '❌'} Max corr to holdings: {n(cr.max_portfolio_corr)}  (threshold: < {CORRELATION_MAX})",
    ]
    if cr.portfolio_correlations:
        for pt, corr in sorted(cr.portfolio_correlations.items(), key=lambda x: -x[1])[:4]:
            flag = "⚠️ " if corr >= CORRELATION_MAX else "✅ "
            lines.append(f"     {flag}{pt}: {corr:.3f}")
    lines += [
        "",
        f"  OVERALL: {analysis.filters_passed}/4 — {analysis.overall_signal}",
    ]
    return "\n".join(lines)


# ─────────────────────────────────────────────
# QUICK TEST
# ─────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    ticker = sys.argv[1] if len(sys.argv) > 1 else "MSFT"
    print(f"\nRunning Dalio analysis on {ticker}...\n")
    result = run_dalio_analysis(ticker)
    print(format_dalio_display(result))
    print()
    print(format_dalio_for_llm(result))
