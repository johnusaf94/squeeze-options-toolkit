"""
investor_roundtable_gui.py  (v2 — Composite Score Edition)
============================================================
Pure Python metrics dashboard + Claude API for Q&A.
No local LLM required. Scores are 100% deterministic.

Requires: pip install yfinance requests openpyxl
"""

import tkinter as tk
from tkinter import scrolledtext, messagebox, filedialog
import threading
import requests
import os
import traceback

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
PORTFOLIO_FILE = "portfolio.xlsx"
MAX_TOKENS     = 512
TEMPERATURE    = 0.3

# ── LOCAL (LM Studio) ──
LM_STUDIO_URL    = "http://localhost:1234/v1/chat/completions"
LM_STUDIO_MODELS = "http://localhost:1234/v1/models"

# ── ONLINE PROVIDERS ──
# Groq — free tier, very fast, OpenAI-compatible API
GROQ_URL     = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODELS  = {
    "Llama 3.3 70B (Fast)":    "llama-3.3-70b-versatile",
    "Llama 3.1 8B (Fastest)":  "llama-3.1-8b-instant",
    "Gemma 3 27B":             "gemma2-9b-it",
    "Mixtral 8x7B":            "mixtral-8x7b-32768",
}

# Together AI — cheap, has 405B
TOGETHER_URL    = "https://api.together.xyz/v1/chat/completions"
TOGETHER_MODELS = {
    "Llama 3.1 405B":       "meta-llama/Meta-Llama-3.1-405B-Instruct-Turbo",
    "Llama 3.3 70B":        "meta-llama/Llama-3.3-70B-Instruct-Turbo",
    "Deepseek V3":          "deepseek-ai/DeepSeek-V3",
}

# Add your API keys here (or set as environment variables)
import os
GROQ_API_KEY    = os.environ.get("GROQ_API_KEY", "")
TOGETHER_API_KEY = os.environ.get("TOGETHER_API_KEY", "")

# Current active backend — toggled by UI switch
# "local" = LM Studio | "groq" = Groq | "together" = Together AI
_ACTIVE_BACKEND = "local"
_ACTIVE_ONLINE_MODEL = list(GROQ_MODELS.values())[0]


def get_active_model() -> str:
    """Get model ID for current backend."""
    if _ACTIVE_BACKEND == "local":
        try:
            resp = requests.get(LM_STUDIO_MODELS, timeout=5)
            if resp.status_code == 200:
                models = resp.json().get("data", [])
                if models:
                    return models[0].get("id", "local-model")
        except Exception:
            pass
        return "local-model"
    return _ACTIVE_ONLINE_MODEL


def get_backend_url() -> str:
    if _ACTIVE_BACKEND == "groq":
        return GROQ_URL
    if _ACTIVE_BACKEND == "together":
        return TOGETHER_URL
    return LM_STUDIO_URL


def get_backend_headers() -> dict:
    headers = {"Content-Type": "application/json"}
    if _ACTIVE_BACKEND == "groq" and GROQ_API_KEY:
        headers["Authorization"] = f"Bearer {GROQ_API_KEY}"
    elif _ACTIVE_BACKEND == "together" and TOGETHER_API_KEY:
        headers["Authorization"] = f"Bearer {TOGETHER_API_KEY}"
    return headers


def check_backend_status() -> tuple:
    """Returns (ok: bool, message: str)"""
    if _ACTIVE_BACKEND == "local":
        try:
            resp = requests.get(LM_STUDIO_MODELS, timeout=5)
            if resp.status_code == 200:
                models = resp.json().get("data", [])
                model_name = models[0].get("id", "unknown") if models else "none loaded"
                return True, f"LM Studio ✓  {model_name}"
            return False, "LM Studio not responding"
        except Exception:
            return False, "LM Studio offline — start it on port 1234"
    elif _ACTIVE_BACKEND == "groq":
        if not GROQ_API_KEY:
            return False, "Groq API key not set — add GROQ_API_KEY env var"
        try:
            resp = requests.get("https://api.groq.com/openai/v1/models",
                                headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                                timeout=8)
            return resp.status_code == 200, f"Groq ✓  {_ACTIVE_ONLINE_MODEL}" if resp.status_code == 200 else "Groq auth failed"
        except Exception as e:
            return False, f"Groq unreachable: {e}"
    elif _ACTIVE_BACKEND == "together":
        if not TOGETHER_API_KEY:
            return False, "Together AI key not set — add TOGETHER_API_KEY env var"
        return True, f"Together AI — {_ACTIVE_ONLINE_MODEL} (key set)"
    return False, "Unknown backend"

# ─────────────────────────────────────────────
# THEME
# ─────────────────────────────────────────────
BG      = "#0A0E14"
BG2     = "#12171F"
BG3     = "#1A2030"
FG      = "#CDD6F4"
FG_DIM  = "#6C7086"
ACCENT  = "#F4C430"
GREEN   = "#A6E3A1"
RED     = "#F38BA8"
YELLOW  = "#F9E2AF"
BLUE    = "#89B4FA"
TEAL    = "#94E2D5"
BORDER  = "#313244"
FONT    = ("Consolas", 10)
FONT_SM = ("Consolas", 9)
FONT_LG = ("Consolas", 12, "bold")
FONT_HD = ("Consolas", 14, "bold")

# ─────────────────────────────────────────────
# ASSET CLASS ROUTING
# ─────────────────────────────────────────────
ASSET_ANALYZERS = {
    # asset_type -> list of (analyzer_key, run_function, args_builder)
    # ETF: Bogle is primary, Druckenmiller for timing, no Buffett moat/DCF
    "ETF": {
        "run":  ["bogle", "druckenmiller"],
        "skip": ["buffett", "weiss_quality", "dalio"],
        "note": "ETF: Bogle cost/diversification + Druckenmiller momentum. Buffett/Weiss blue chip criteria not applicable.",
    },
    # REIT: Weiss yield is primary, Buffett modified (no moat score), Dalio debt
    "REIT": {
        "run":  ["buffett", "weiss", "bogle", "dalio", "druckenmiller"],
        "skip": [],
        "note": "REIT: Full analysis. Note — Buffett DCF unreliable for REITs (use FFO not FCF). Weiss yield method ideal.",
    },
    # Regular stock: run everything
    "STOCK": {
        "run":  ["buffett", "weiss", "bogle", "dalio", "druckenmiller"],
        "skip": [],
        "note": "",
    },
    # Mutual fund: same as ETF
    "MUTUAL_FUND": {
        "run":  ["bogle", "druckenmiller"],
        "skip": ["buffett", "weiss", "dalio"],
        "note": "Mutual Fund: Bogle cost analysis primary. Individual stock metrics not applicable.",
    },
    "UNKNOWN": {
        "run":  ["buffett", "weiss", "bogle", "dalio", "druckenmiller"],
        "skip": [],
        "note": "Unknown asset type — running all analyzers with caveats.",
    },
}

def detect_asset_class(info: dict) -> str:
    """Detect asset type from yfinance info dict."""
    qt = (info.get("quoteType") or "").upper()
    if qt == "ETF":
        return "ETF"
    if qt == "MUTUALFUND":
        return "MUTUAL_FUND"
    sector = info.get("sector") or ""
    industry = (info.get("industry") or "").upper()
    name = (info.get("longName") or "").upper()
    if qt == "EQUITY":
        if sector == "Real Estate" or "REIT" in name or "REAL ESTATE INV" in name:
            return "REIT"
        return "STOCK"
    if info.get("fundFamily"):
        return "ETF"
    return "UNKNOWN"


CLIENT_CONTEXT = """JOHNATHAN RUSH — CLIENT PROFILE
Age 30 | Port Orange FL | Married, one daughter | 90% disabled veteran
VA disability = permanent inflation-adjusted income floor for life.
INVESTMENT PHASE: Aggressive accumulation. He does NOT need this portfolio to survive.
PRIMARY GOAL: Maximize total real return over 30 years.
RISK TOLERANCE: High — VA floor absorbs worst-case drawdown scenarios.
HOLDINGS: Roth 401k (O, MSFT, VICI) | Taxable (PFE, GIS, UPS) | $9k cash to deploy.
DO NOT default to defensive income stocks. Evaluate on quality and total return."""


# ─────────────────────────────────────────────
# PORTFOLIO LOADER
# ─────────────────────────────────────────────
def load_portfolio_context(filepath):
    if not os.path.exists(filepath):
        return ""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sections = []
        if "Holdings" in wb.sheetnames:
            ws = wb["Holdings"]
            headers = []
            lines = []
            for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
                vals = [str(v).strip() if v is not None else "" for v in row]
                if i == 0:
                    headers = vals
                    continue
                if not vals[0] or vals[0] in ("", "TOTALS"):
                    continue
                line = ", ".join(
                    f"{headers[j]}: {vals[j]}"
                    for j in range(min(len(headers), len(vals)))
                    if vals[j] and headers[j]
                )
                if line:
                    lines.append(f"  - {line}")
            if lines:
                sections.append("HOLDINGS:\n" + "\n".join(lines))
        return "\n\n".join(sections)
    except Exception as e:
        return f"[Portfolio load error: {e}]"


def fetch_live_prices(filepath):
    try:
        import openpyxl, yfinance as yf
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Holdings"]
        tickers = []
        row_map = {}
        for row in ws.iter_rows(min_row=5, values_only=False):
            t = row[0].value
            if not t or str(t).strip() in ("", "TOTALS", "CASH"):
                continue
            t = str(t).strip().upper()
            tickers.append(t)
            row_map[t] = row[4].row
        if tickers:
            data = yf.download(tickers, period="1d", progress=False, auto_adjust=True)
            close = data["Close"]
            thin = Side(style="thin", color="30363D")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for t in tickers:
                try:
                    price = float(close[t].iloc[-1]) if len(tickers) > 1 else float(close.iloc[-1])
                    c = ws.cell(row=row_map[t], column=5, value=round(price, 4))
                    c.font = Font(name="Consolas", size=10, color="007700", bold=True)
                    c.fill = PatternFill("solid", start_color="E8FFE8")
                    c.number_format = "$#,##0.00"
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.border = border
                except Exception:
                    pass
        wb.save(filepath)
    except Exception:
        pass


# ─────────────────────────────────────────────
# ANALYSIS PIPELINE
# ─────────────────────────────────────────────
def run_full_analysis(ticker, portfolio_path=PORTFOLIO_FILE):
    import yfinance as yf
    from ticker_resolver import resolve_ticker, fetch_live_data
    from composite_score import build_composite
    from buffett_analyzer import run_buffett_analysis, fetch_buffett_indicator
    from weiss_analyzer import run_weiss_analysis
    from bogle_analyzer import run_bogle_analysis
    from dalio_analyzer import run_dalio_analysis

    resolved, company_name, ok = resolve_ticker(ticker)
    if not ok:
        return None, f"Could not resolve ticker '{ticker}' — try the exact ticker symbol."

    results = {"ticker": resolved, "company_name": company_name}
    live_data = fetch_live_data(resolved)
    results["live_data"] = live_data

    from druckenmiller_analyzer import run_druckenmiller_analysis
    # Detect asset class and route to appropriate analyzers
    try:
        asset_class = detect_asset_class(yf.Ticker(resolved).info)
    except Exception:
        asset_class = "UNKNOWN"

    routing = ASSET_ANALYZERS.get(asset_class, ASSET_ANALYZERS["UNKNOWN"])
    results["asset_class"] = asset_class
    results["asset_note"]  = routing["note"]
    skip = set(routing["skip"])

    all_analyzers = [
        ("buffett",       run_buffett_analysis,       (resolved,)),
        ("weiss",         run_weiss_analysis,         (resolved,)),
        ("bogle",         run_bogle_analysis,         (resolved, portfolio_path)),
        ("dalio",         run_dalio_analysis,         (resolved, portfolio_path)),
        ("druckenmiller", run_druckenmiller_analysis, (resolved,)),
    ]

    for name, fn, args in all_analyzers:
        # Skip analyzers not appropriate for this asset class
        base = name.split("_")[0]
        if base in skip or name in skip:
            results[name] = None
            results[f"{name}_skipped"] = True
            continue
        try:
            results[name] = fn(*args)
        except Exception as e:
            results[name] = None
            results[f"{name}_error"] = str(e)

    try:
        bi = fetch_buffett_indicator()
        market_ctx = f"Buffett Indicator {bi.ratio*100:.0f}% — {bi.signal}" if bi.ratio else ""
    except Exception:
        market_ctx = ""

    # Build the skipped set from routing + live data checks
    active_skipped = set(routing.get("skip", []))

    # Auto-skip Weiss yield if no dividend (non-dividend stocks)
    if live_data and (not live_data.dividend_rate or live_data.dividend_rate == 0):
        active_skipped.add("weiss_yield")

    composite = build_composite(
        ticker=resolved,
        company_name=company_name,
        buffett_analysis=results.get("buffett"),
        weiss_analysis=results.get("weiss"),
        bogle_analysis=results.get("bogle"),
        dalio_analysis=results.get("dalio"),
        druckenmiller_analysis=results.get("druckenmiller"),
        live_data=live_data,
        market_context_str=market_ctx,
        skipped=active_skipped,
    )
    results["composite"] = composite
    return results, None


# ─────────────────────────────────────────────
# CLAUDE API
# ─────────────────────────────────────────────
def ask_lm_studio(question, composite_context, portfolio_context):
    """Send Q&A to current active backend (local or online)."""
    import re

    system = (
        f"You are a financial analyst reviewing a pre-calculated composite investment score.\n"
        f"{CLIENT_CONTEXT}\n\n"
        f"Scores computed by Python pipeline:\n{composite_context}\n\n"
        f"Portfolio:\n{portfolio_context}\n\n"
        f"RULES: Answer in plain English only. No LaTeX, no markdown symbols. "
        f"Reference the actual score numbers. Max 200 words. No generic advice."
    )

    payload = {
        "model":       get_active_model(),
        "messages":    [
            {"role": "system", "content": system},
            {"role": "user",   "content": question},
        ],
        "max_tokens":  MAX_TOKENS,
        "temperature": TEMPERATURE,
        "stream":      False,
    }

    # Local-only params
    if _ACTIVE_BACKEND == "local":
        payload["n_ctx"] = 8192

    timeout = 300 if _ACTIVE_BACKEND == "local" else 60

    try:
        resp = requests.post(
            get_backend_url(),
            json=payload,
            headers=get_backend_headers(),
            timeout=timeout,
        )
        resp.raise_for_status()
        raw = resp.json()["choices"][0]["message"]["content"].strip()

        # Strip DeepSeek/R1 think blocks
        if "<think>" in raw:
            end = raw.find("</think>")
            if end != -1:
                raw = raw[end + 8:].strip()

        # Strip LaTeX artifacts
        raw = re.sub(r"\\text\{([^}]*)\}", r"\1", raw)
        raw = re.sub(r"\$([A-Za-z0-9.+\-_ ]{1,30})\$", r"\1", raw)
        raw = re.sub(r"\*\*([^*]+)\*\*", r"\1", raw)
        raw = re.sub(r"^#{1,3}\s+", "", raw, flags=re.MULTILINE)
        return raw

    except requests.exceptions.ConnectionError:
        if _ACTIVE_BACKEND == "local":
            return "❌ Cannot connect to LM Studio on port 1234. Make sure it is running."
        return f"❌ Cannot reach {_ACTIVE_BACKEND} API. Check your internet connection."
    except requests.exceptions.Timeout:
        return "❌ Request timed out. Try a smaller/faster model."
    except Exception as e:
        err = str(e)
        if "401" in err or "403" in err:
            return f"❌ API key error for {_ACTIVE_BACKEND}. Check your key in the settings panel."
        if "400" in err and _ACTIVE_BACKEND == "local":
            return "❌ 400 error — set Context Length to 8192 in LM Studio server tab."
        return f"❌ {_ACTIVE_BACKEND} error: {err}"

ask_claude = ask_lm_studio


# ─────────────────────────────────────────────
# MAIN APPLICATION
# ─────────────────────────────────────────────
class RoundtableApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Investor Roundtable")
        self.root.geometry("1300x900")
        self.root.configure(bg=BG)
        self.root.minsize(1100, 720)

        self.is_running      = False
        self.stop_requested  = False
        self.session_results = None
        self.portfolio_ctx   = load_portfolio_context(PORTFOLIO_FILE)
        self.mode_var        = tk.StringVar(value='composite')  # 'composite' or 'long'

        self._build_ui()
        threading.Thread(target=self._init_prices, daemon=True).start()

    def _on_backend_change(self, selection):
        global _ACTIVE_BACKEND, _ACTIVE_ONLINE_MODEL, GROQ_API_KEY, TOGETHER_API_KEY
        backend, model_id = self._model_options.get(selection, ("local", "local-model"))
        _ACTIVE_BACKEND = backend
        _ACTIVE_ONLINE_MODEL = model_id

        # Show/hide API key field
        for w in self._apikey_frame.winfo_children():
            w.pack_forget()

        if backend != "local":
            existing_key = GROQ_API_KEY if backend == "groq" else TOGETHER_API_KEY
            self._apikey_var.set(existing_key or "")
            self._apikey_lbl.config(text=f"{backend.capitalize()} Key:")
            self._apikey_lbl.pack(side="left")
            self._apikey_entry.pack(side="left", padx=2)
            self._apikey_btn.pack(side="left")

        threading.Thread(target=self._check_backend, daemon=True).start()

    def _set_api_key(self):
        global GROQ_API_KEY, TOGETHER_API_KEY
        key = self._apikey_var.get().strip()
        if _ACTIVE_BACKEND == "groq":
            GROQ_API_KEY = key
        elif _ACTIVE_BACKEND == "together":
            TOGETHER_API_KEY = key
        threading.Thread(target=self._check_backend, daemon=True).start()

    def _check_backend(self):
        self.conn_lbl.config(text="⏳ checking...", fg=FG_DIM)
        ok, msg = check_backend_status()
        color = GREEN if ok else RED
        icon  = "🟢" if ok else "🔴"
        self.conn_lbl.config(text=f"{icon} {msg}", fg=color)


    def _build_portfolio_tab(self):
        """Portfolio Builder — Michael's recursive discovery engine."""
        import tkinter.ttk as ttk
        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            HAS_MATPLOTLIB = True
        except ImportError:
            HAS_MATPLOTLIB = False

        parent = self.tab_portfolio
        self._pb_running   = False
        self._pb_stop      = False
        self._pb_portfolio = None
        self._pb_candidates = []

        # ── LEFT: Scrollable Controls ────────────────────────────────────
        # Wrap in a canvas so it scrolls when content exceeds window height
        ctrl_outer = tk.Frame(parent, bg=BG2, width=290)
        ctrl_outer.pack(side="left", fill="y")
        ctrl_outer.pack_propagate(False)

        ctrl_canvas = tk.Canvas(ctrl_outer, bg=BG2, width=270,
                                highlightthickness=0, bd=0)
        ctrl_scrollbar = tk.Scrollbar(ctrl_outer, orient="vertical",
                                       command=ctrl_canvas.yview)
        ctrl_canvas.configure(yscrollcommand=ctrl_scrollbar.set)

        ctrl_scrollbar.pack(side="right", fill="y")
        ctrl_canvas.pack(side="left", fill="both", expand=True)

        ctrl = tk.Frame(ctrl_canvas, bg=BG2)
        ctrl_window = ctrl_canvas.create_window((0, 0), window=ctrl, anchor="nw")

        def _on_ctrl_configure(e):
            ctrl_canvas.configure(scrollregion=ctrl_canvas.bbox("all"))
            ctrl_canvas.itemconfig(ctrl_window, width=ctrl_canvas.winfo_width())

        ctrl.bind("<Configure>", _on_ctrl_configure)
        ctrl_canvas.bind("<Configure>", lambda e: ctrl_canvas.itemconfig(
            ctrl_window, width=e.width))

        # Mouse wheel scrolling
        def _on_mousewheel(e):
            ctrl_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        ctrl_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        tk.Label(ctrl, text="PORTFOLIO BUILDER", font=FONT_LG,
                 bg=BG2, fg=ACCENT).pack(pady=(14,4), padx=14, anchor="w")
        tk.Label(ctrl, text="Portfolio Builder — Composite Score Engine",
                 font=FONT_SM, bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=10)

        # ── Inputs ──
        def labeled_entry(parent, label, default, is_int=False):
            f = tk.Frame(parent, bg=BG2)
            f.pack(fill="x", padx=14, pady=3)
            tk.Label(f, text=label, font=FONT_SM, bg=BG2, fg=FG_DIM,
                     width=18, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(default))
            tk.Entry(f, textvariable=var, font=FONT_SM, bg=BG3, fg=FG,
                     insertbackground=FG, relief="flat", bd=4, width=10).pack(side="left")
            return var

        self._pb_cash   = labeled_entry(ctrl, "Starting Cash ($)",     "10000")
        self._pb_years  = labeled_entry(ctrl, "Years to Retirement",   "30")
        self._pb_monthly= labeled_entry(ctrl, "Monthly Contrib ($)",   "600")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Sector selector ──
        tk.Label(ctrl, text="SECTORS TO SCAN", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        from portfolio_builder import SECTOR_UNIVERSE
        self._pb_sectors = {}
        sectors_frame = tk.Frame(ctrl, bg=BG2)
        sectors_frame.pack(fill="x", padx=10)

        all_sectors = list(SECTOR_UNIVERSE.keys())
        # Default select most growth-oriented sectors
        default_on = {"Technology","Healthcare","Consumer Discretionary",
                      "Financials","Industrials","Communication Services"}

        for i, sec in enumerate(all_sectors):
            var = tk.BooleanVar(value=sec in default_on)
            self._pb_sectors[sec] = var
            short = sec[:22]
            tk.Checkbutton(sectors_frame, text=short, variable=var,
                           font=FONT_SM, bg=BG2, fg=FG, selectcolor=BG3,
                           activebackground=BG2, relief="flat").pack(anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Agent selector ──
        tk.Label(ctrl, text="ANALYST AGENTS", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        self._pb_agents = {}
        agents = [
            ("buffett",        "🎩 Buffett"),
            ("weiss",          "📈 Weiss"),
            ("bogle",          "📊 Bogle"),
            ("dalio",          "🌊 Dalio"),
            ("druckenmiller",  "📡 Druckenmiller"),
        ]
        agents_frame = tk.Frame(ctrl, bg=BG2)
        agents_frame.pack(fill="x", padx=10)
        for key, label in agents:
            var = tk.BooleanVar(value=True)
            self._pb_agents[key] = var
            tk.Checkbutton(agents_frame, text=label, variable=var,
                           font=FONT_SM, bg=BG2, fg=FG, selectcolor=BG3,
                           activebackground=BG2, relief="flat").pack(anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Portfolio Constraints ──
        tk.Label(ctrl, text="PORTFOLIO CONSTRAINTS", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        # Tickers per sector
        f_tps = tk.Frame(ctrl, bg=BG2)
        f_tps.pack(fill="x", padx=14, pady=2)
        tk.Label(f_tps, text="Tickers/sector", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_tickers_per = tk.StringVar(value="5")
        tk.Entry(f_tps, textvariable=self._pb_tickers_per, font=FONT_SM,
                 bg=BG3, fg=FG, relief="flat", bd=4, width=6).pack(side="left")

        # Max sector allocation slider
        f_sec = tk.Frame(ctrl, bg=BG2)
        f_sec.pack(fill="x", padx=14, pady=4)
        tk.Label(f_sec, text="Max Sector %", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_max_sector = tk.IntVar(value=35)
        self._pb_sector_lbl = tk.Label(f_sec, text="35%", font=FONT_SM,
                                        bg=BG2, fg=ACCENT, width=5)
        self._pb_sector_lbl.pack(side="right")

        def _update_sector_lbl(val):
            self._pb_sector_lbl.config(text=f"{int(float(val))}%")

        tk.Scale(ctrl, from_=10, to=100, orient="horizontal",
                 variable=self._pb_max_sector,
                 command=_update_sector_lbl,
                 bg=BG2, fg=FG, troughcolor=BG3,
                 highlightthickness=0, bd=0,
                 sliderlength=16, showvalue=False,
                 length=220).pack(padx=14, pady=(0,4))

        # Target return slider
        f_ret = tk.Frame(ctrl, bg=BG2)
        f_ret.pack(fill="x", padx=14, pady=2)
        tk.Label(f_ret, text="Target Return %", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_target_return = tk.IntVar(value=15)
        self._pb_return_lbl = tk.Label(f_ret, text="15%", font=FONT_SM,
                                        bg=BG2, fg=GREEN, width=5)
        self._pb_return_lbl.pack(side="right")

        def _update_return_lbl(val):
            v = int(float(val))
            color = GREEN if v <= 15 else (YELLOW if v <= 20 else RED)
            self._pb_return_lbl.config(text=f"{v}%", fg=color)

        tk.Scale(ctrl, from_=5, to=30, orient="horizontal",
                 variable=self._pb_target_return,
                 command=_update_return_lbl,
                 bg=BG2, fg=FG, troughcolor=BG3,
                 highlightthickness=0, bd=0,
                 sliderlength=16, showvalue=False,
                 length=220).pack(padx=14, pady=(0,6))

        # Max position per stock slider
        f_pos = tk.Frame(ctrl, bg=BG2)
        f_pos.pack(fill="x", padx=14, pady=4)
        tk.Label(f_pos, text="Max Position %", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_max_position = tk.IntVar(value=20)
        self._pb_position_lbl = tk.Label(f_pos, text="20%", font=FONT_SM,
                                          bg=BG2, fg=ACCENT, width=5)
        self._pb_position_lbl.pack(side="right")

        def _update_position_lbl(val):
            self._pb_position_lbl.config(text=f"{int(float(val))}%")

        tk.Scale(ctrl, from_=5, to=50, orient="horizontal",
                 variable=self._pb_max_position,
                 command=_update_position_lbl,
                 bg=BG2, fg=FG, troughcolor=BG3,
                 highlightthickness=0, bd=0,
                 sliderlength=16, showvalue=False,
                 length=220).pack(padx=14, pady=(0,6))

        tk.Label(ctrl, text="5% = highly diversified  |  50% = concentrated",
                 font=("Consolas",7), bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w")

        tk.Label(ctrl, text="5% = conservative  |  30% = aggressive",
                 font=("Consolas",7), bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Include ETFs toggle ──
        self._pb_include_etf = tk.BooleanVar(value=False)
        etf_frame = tk.Frame(ctrl, bg=BG2)
        etf_frame.pack(fill="x", padx=10, pady=2)

        self._pb_etf_btn = tk.Button(
            etf_frame,
            text="☐  Include Sector ETFs",
            font=FONT_SM,
            bg=BG3, fg=FG_DIM,
            relief="flat", cursor="hand2",
            anchor="w", padx=10, pady=5,
            command=self._pb_toggle_etf,
        )
        self._pb_etf_btn.pack(fill="x")
        tk.Label(ctrl, text="Adds top ETFs by market cap per sector",
                 font=("Consolas",7), bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        # Run button
        self._pb_run_btn = tk.Button(ctrl, text="▶  Build Portfolio",
                                      font=("Consolas",11,"bold"),
                                      bg=ACCENT, fg="#000000",
                                      relief="flat", cursor="hand2",
                                      padx=14, pady=6,
                                      command=self._pb_toggle)
        self._pb_run_btn.pack(fill="x", padx=10, pady=8)

        self._pb_status = tk.Label(ctrl, text="Ready", font=FONT_SM,
                                    bg=BG2, fg=FG_DIM, wraplength=240)
        self._pb_status.pack(padx=14, anchor="w")

        # ── RIGHT: Output split ──────────────────────────────────────────
        right = tk.Frame(parent, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Top: pie chart area
        chart_frame = tk.Frame(right, bg=BG, height=320)
        chart_frame.pack(fill="x", pady=0)
        chart_frame.pack_propagate(False)

        if HAS_MATPLOTLIB:
            fig = Figure(figsize=(5, 3.2), dpi=90, facecolor=BG)
            fig.subplots_adjust(left=0.02, right=0.98, top=0.92, bottom=0.02)
            self._pb_fig = fig
            self._pb_ax  = fig.add_subplot(111)
            self._pb_ax.set_facecolor(BG)
            self._pb_canvas = FigureCanvasTkAgg(fig, master=chart_frame)
            self._pb_canvas.get_tk_widget().pack(fill="both", expand=True)
            self._draw_empty_pie()
        else:
            self._pb_canvas = None
            self._pb_fig    = None
            tk.Label(chart_frame,
                     text="Install matplotlib for pie chart: pip install matplotlib",
                     font=FONT_SM, bg=BG, fg=YELLOW).pack(pady=40)

        # Bottom: log + results
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x")
        self._pb_log = scrolledtext.ScrolledText(
            right, wrap="word", font=FONT_SM, bg=BG, fg=FG,
            insertbackground=FG, relief="flat", borderwidth=0,
            state="disabled", padx=16, pady=10,
        )
        self._pb_log.pack(fill="both", expand=True)

        # Log tags
        for tag, cfg in [
            ("header",  {"font": FONT_LG,               "foreground": ACCENT}),
            ("dim",     {                                 "foreground": FG_DIM}),
            ("green",   {                                 "foreground": GREEN}),
            ("red",     {                                 "foreground": RED}),
            ("yellow",  {                                 "foreground": YELLOW}),
            ("blue",    {                                 "foreground": BLUE}),
            ("accent_teal", {"font": ("Consolas",10,"bold"), "foreground": TEAL}),
            ("include", {"font": ("Consolas",9,"bold"),  "foreground": GREEN}),
            ("reject",  {"font": ("Consolas",9),         "foreground": RED}),
            ("watch",   {"font": ("Consolas",9),         "foreground": YELLOW}),
        ]:
            self._pb_log.tag_config(tag, **cfg)

    def _draw_empty_pie(self):
        if not self._pb_fig:
            return
        ax = self._pb_ax
        ax.clear()
        ax.set_facecolor(BG)
        ax.pie([1], colors=[BG3], startangle=90)
        ax.set_title("Waiting for portfolio...", color=FG_DIM,
                     fontsize=9, fontfamily="monospace")
        self._pb_canvas.draw()

    def _update_pie(self, portfolio):
        if not self._pb_fig or not portfolio or not portfolio.positions:
            return
        ax = self._pb_ax
        ax.clear()
        ax.set_facecolor(BG)

        labels  = [f"{p.ticker}\n{p.allocation_pct:.1f}%" for p in portfolio.positions]
        sizes   = [p.allocation_pct for p in portfolio.positions]
        if portfolio.cash_remaining > 0:
            cash_pct = portfolio.cash_remaining / portfolio.total_value * 100
            if cash_pct > 0.5:
                labels.append(f"CASH\n{cash_pct:.1f}%")
                sizes.append(cash_pct)

        # Vivid color palette cycling
        palette = [
            "#F4C430","#89B4FA","#A6E3A1","#F38BA8","#94E2D5",
            "#CBA6F7","#FAB387","#A6ADC8","#74C7EC","#EBA0AC",
            "#B4BEFE","#F9E2AF","#89DCEB","#45475A","#6C7086",
        ]
        colors = [palette[i % len(palette)] for i in range(len(sizes))]

        wedges, texts = ax.pie(
            sizes, labels=None, colors=colors,
            startangle=90, wedgeprops={"linewidth": 1.5, "edgecolor": BG}
        )

        # Clean legend
        ax.legend(wedges, labels, loc="center left", bbox_to_anchor=(1, 0.5),
                  fontsize=7, frameon=False,
                  labelcolor=FG, facecolor=BG)

        ret_str = f"{portfolio.expected_return:.1%}" if portfolio.expected_return else "N/A"
        sharpe_str = f"{portfolio.expected_sharpe:.2f}" if portfolio.expected_sharpe else "N/A"
        target = "✅" if portfolio.meets_target else "⚠️"
        ax.set_title(
            f"{target} {len(portfolio.positions)} positions | "
            f"Exp return: {ret_str} | Sharpe: {sharpe_str}",
            color=FG, fontsize=8, fontfamily="monospace", pad=8
        )
        self._pb_canvas.draw()

    def _pb_log_write(self, text, tag=None):
        self._pb_log.config(state="normal")
        self._pb_log.insert("end", text, tag) if tag else self._pb_log.insert("end", text)
        self._pb_log.see("end")
        self._pb_log.config(state="disabled")
        self.root.update_idletasks()

    def _pb_toggle_etf(self):
        """Toggle Include ETFs button state."""
        current = self._pb_include_etf.get()
        self._pb_include_etf.set(not current)
        if self._pb_include_etf.get():
            self._pb_etf_btn.config(
                text="☑  Include Sector ETFs",
                bg="#238636", fg="#FFFFFF"
            )
        else:
            self._pb_etf_btn.config(
                text="☐  Include Sector ETFs",
                bg=BG3, fg=FG_DIM
            )

    def _pb_toggle(self):
        if self._pb_running:
            self._pb_stop = True
            self._pb_run_btn.config(text="⏹ Stopping...", bg=RED, state="disabled")
        else:
            self._pb_start()

    def _pb_start(self):
        try:
            cash       = float(self._pb_cash.get().replace(",",""))
            years      = int(self._pb_years.get())
            monthly    = float(self._pb_monthly.get().replace(",",""))
            tps        = int(self._pb_tickers_per.get())
            max_sector   = self._pb_max_sector.get()   / 100.0
            max_position = self._pb_max_position.get() / 100.0
            target_ret   = self._pb_target_return.get() / 100.0
            include_etf  = self._pb_include_etf.get()
        except ValueError:
            self._pb_log_write("❌ Invalid inputs — check Starting Cash, Years, Monthly Contrib.\n", "red")
            return

        sectors = [s for s, v in self._pb_sectors.items() if v.get()]
        agents  = [a for a, v in self._pb_agents.items() if v.get()]

        if not sectors:
            self._pb_log_write("❌ Select at least one sector.\n", "red")
            return
        if not agents:
            self._pb_log_write("❌ Select at least one analyst agent.\n", "red")
            return

        self._pb_running = True
        self._pb_stop    = False
        self._pb_run_btn.config(bg=RED, fg="#000000", text="⏹ Stop")
        self._draw_empty_pie()

        threading.Thread(
            target=self._pb_thread,
            args=(cash, years, monthly, sectors, agents, tps,
                  max_sector, max_position, target_ret, include_etf),
            daemon=True,
        ).start()

    def _pb_thread(self, cash, years, monthly, sectors, agents, tps,
                    max_sector=0.35, max_position=0.20, target_ret=0.15, include_etf=False):
        try:
            self._pb_log.config(state="normal")
            self._pb_log.delete("1.0", "end")
            self._pb_log.config(state="disabled")

            self._pb_log_write("\n")
            self._pb_log_write("  PORTFOLIO BUILDER — DISCOVERY RUN\n", "header")
            self._pb_log_write(f"  Cash: ${cash:,.0f}  |  Horizon: {years}yr  |  Contrib: ${monthly:,.0f}/mo\n", "dim")
            self._pb_log_write(f"  Target Return: {target_ret:.0%}  |  Max Sector: {max_sector:.0%}  |  Max Position: {max_position:.0%}  |  ETFs: {'YES' if include_etf else 'NO'}\n", "dim")
            self._pb_log_write(f"  Scanning {len(sectors)} sectors × {tps} tickers = "
                                f"up to {len(sectors)*tps} candidates\n", "dim")
            self._pb_log_write(f"  Agents: {', '.join(agents)}\n\n", "dim")

            self._pb_status.config(text="Scanning sectors...")

            from portfolio_builder import run_portfolio_builder

            def progress(ticker, msg):
                if self._pb_stop:
                    return
                # Color by score label (no binary gate anymore)
                tag = "dim"
                if "STRONG" in msg:
                    tag = "include"
                elif "GOOD" in msg:
                    tag = "green"
                elif "FAIR" in msg:
                    tag = "yellow"
                elif "WEAK" in msg:
                    tag = "reject"
                elif "⏭" in msg:
                    tag = "dim"
                self._pb_log_write(f"  {msg}\n", tag)
                self._pb_status.config(text=f"Scanning {ticker}...")

            portfolio, candidates = run_portfolio_builder(
                starting_cash=cash,
                years=years,
                monthly_contrib=monthly,
                selected_sectors=sectors,
                selected_agents=agents,
                tickers_per_sector=tps,
                max_sector_pct=max_sector,
                max_position_pct=max_position,
                target_annual_return=target_ret,
                include_etf=include_etf,
                progress_callback=progress,
            )

            if self._pb_stop:
                self._pb_log_write("\n  ⏹ Stopped.\n", "dim")
                return

            self._pb_portfolio  = portfolio
            self._pb_candidates = candidates

            # ── Michael's results ──
            self._pb_log_write("\n")
            self._pb_log_write("  ─" * 30 + "\n", "dim")
            self._pb_log_write("  PORTFOLIO BUILDER RESULTS\n", "header")
            self._pb_log_write("  ─" * 30 + "\n\n", "dim")

            self._pb_log_write(f"  {portfolio.pb_summary}\n\n", "include")

            # Position table
            self._pb_log_write(
                f"  {'Ticker':<8} {'Company':<24} {'Sector':<22} {'Alloc':>6}  "
                f"{'Return':>7}  {'Sharpe':>6}  {'Vol':>7}  {'VolScore':>8}  {'Comp':>5}\n", "blue"
            )
            self._pb_log_write(f"  {'─'*92}\n", "dim")

            # Pull volatility from candidates for display
            cand_map = {c.ticker: c for c in candidates}
            for pos in portfolio.positions:
                ret_s  = f"{pos.annual_return:.1%}"  if pos.annual_return  else "N/A"
                sha_s  = f"{pos.sharpe_ratio:.2f}"   if pos.sharpe_ratio   else "N/A"
                cand   = cand_map.get(pos.ticker)
                vol_s  = f"{cand.annual_volatility:.1%}" if cand and cand.annual_volatility else "N/A"
                pb_s   = f"{cand.pb_score:.0f}"      if cand else "N/A"
                # Color by pb_score
                tag = "green" if (cand and cand.pb_score >= 65) else ("yellow" if (cand and cand.pb_score >= 45) else "dim")
                self._pb_log_write(
                    f"  {pos.ticker:<8} {pos.company_name[:24]:<24} "
                    f"{pos.sector[:22]:<22} {pos.allocation_pct:>5.1f}%  "
                    f"{ret_s:>7}  {sha_s:>6}  {vol_s:>7}  {pb_s:>8}  {pos.composite_score:>5.1f}\n",
                    tag
                )

            self._pb_log_write(f"\n  Cash remaining: ${portfolio.cash_remaining:,.0f}\n", "dim")

            # Sector breakdown
            self._pb_log_write("\n  SECTOR ALLOCATION\n", "blue")
            for sec, pct in sorted(portfolio.sector_weights.items(), key=lambda x: -x[1]):
                bar = "█" * int(pct / 3) + "░" * (33 - int(pct / 3))
                self._pb_log_write(f"  {sec:<26} {pct:>5.1f}%  {bar}\n", "dim")

            # Projection — three scenarios
            if hasattr(portfolio, "projected_value_at_retirement") and portfolio.projected_value_at_retirement:
                self._pb_log_write(f"\n  {'─'*60}\n", "dim")
                self._pb_log_write(f"  PROJECTED VALUE AT YEAR {years}\n", "header")
                self._pb_log_write(
                    f"  Starting: ${cash:,.0f}  |  "
                    f"Monthly: ${monthly:,.0f}  |  "
                    f"Horizon: {years} years\n", "dim"
                )
                if hasattr(portfolio, "return_was_capped") and portfolio.return_was_capped:
                    self._pb_log_write(
                        f"  ⚠️  Raw portfolio return was {portfolio.expected_return:.1%} — "
                        f"capped at 20% for projection realism.\n"
                        f"  Historical data often overstates future returns.\n", "yellow"
                    )
                self._pb_log_write("\n", "dim")
                for label, val in portfolio.projection_scenarios.items():
                    self._pb_log_write(f"  {label:<40} ${val:>16,.0f}\n", "include")
                self._pb_log_write("\n", "dim")
                self._pb_log_write(
                    "  Note: These are projections, not guarantees. Real returns vary.\n"
                    "  The conservative 7% scenario reflects long-run S&P 500 average.\n", "dim"
                )

            self._pb_log_write("\n")

            # Update pie chart
            self._update_pie(portfolio)
            self._pb_status.config(text=f"Done — {len(portfolio.positions)} positions selected")

        except Exception as e:
            import traceback
            self._pb_log_write(f"\n❌ Error: {e}\n", "red")
            self._pb_log_write(traceback.format_exc(), "dim")
            self._pb_status.config(text="Error")
        finally:
            self._pb_running = False
            self._pb_stop    = False
            self._pb_run_btn.config(state="normal", text="▶  Build Portfolio",
                                     bg=ACCENT, fg="#000000")


    def _build_squeeze_tab(self):
        """Squeeze Searcher — scans S&P500 from smallest to largest market cap."""
        parent = self.tab_squeeze
        self._sq_running   = False
        self._sq_stop      = False
        self._sq_results   = []   # list of (combined_score, gill, chamath, ticker)

        # ── LEFT: Controls ──────────────────────────────────────────────
        sq_outer = tk.Frame(parent, bg=BG2, width=270)
        sq_outer.pack(side="left", fill="y")
        sq_outer.pack_propagate(False)

        sq_canvas = tk.Canvas(sq_outer, bg=BG2, width=250,
                               highlightthickness=0, bd=0)
        sq_scroll = tk.Scrollbar(sq_outer, orient="vertical",
                                  command=sq_canvas.yview)
        sq_canvas.configure(yscrollcommand=sq_scroll.set)
        sq_scroll.pack(side="right", fill="y")
        sq_canvas.pack(side="left", fill="both", expand=True)
        ctrl = tk.Frame(sq_canvas, bg=BG2)
        cw = sq_canvas.create_window((0, 0), window=ctrl, anchor="nw")
        ctrl.bind("<Configure>", lambda e: sq_canvas.configure(
            scrollregion=sq_canvas.bbox("all")))
        sq_canvas.bind("<Configure>", lambda e: sq_canvas.itemconfig(cw, width=e.width))
        sq_canvas.bind_all("<MouseWheel>", lambda e: sq_canvas.yview_scroll(
            int(-1*(e.delta/120)), "units"))

        tk.Label(ctrl, text="SQUEEZE SEARCHER", font=FONT_LG,
                 bg=BG2, fg=ACCENT).pack(pady=(14,2), padx=14, anchor="w")
        tk.Label(ctrl, text="S&P500 • Smallest → Largest Market Cap",
                 font=FONT_SM, bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=10)

        # Search scope
        tk.Label(ctrl, text="SEARCH SCOPE", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        def labeled_sq(label, default):
            f = tk.Frame(ctrl, bg=BG2)
            f.pack(fill="x", padx=14, pady=2)
            tk.Label(f, text=label, font=FONT_SM, bg=BG2, fg=FG_DIM,
                     width=16, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(default))
            tk.Entry(f, textvariable=var, font=FONT_SM, bg=BG3, fg=FG,
                     insertbackground=FG, relief="flat", bd=4, width=8).pack(side="left")
            return var

        self._sq_max_stocks  = labeled_sq("Max stocks", "500")
        self._sq_top_results = labeled_sq("Show top N", "20")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # Squeeze thresholds
        tk.Label(ctrl, text="MINIMUM THRESHOLDS", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        self._sq_min_si  = labeled_sq("Min Short Int %", "5")
        self._sq_min_dtc = labeled_sq("Min Days to Cover", "1")
        self._sq_min_score = labeled_sq("Min Combined Score", "30")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # Agent selector
        tk.Label(ctrl, text="SQUEEZE AGENTS", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        self._sq_use_gill    = tk.BooleanVar(value=True)
        self._sq_use_chamath = tk.BooleanVar(value=True)
        af = tk.Frame(ctrl, bg=BG2)
        af.pack(fill="x", padx=10)
        tk.Checkbutton(af, text="🎮 Keith Gill (DFV)",
                       variable=self._sq_use_gill,
                       font=FONT_SM, bg=BG2, fg=FG, selectcolor=BG3,
                       activebackground=BG2, relief="flat").pack(anchor="w")
        tk.Checkbutton(af, text="💰 Chamath Palihapitiya",
                       variable=self._sq_use_chamath,
                       font=FONT_SM, bg=BG2, fg=FG, selectcolor=BG3,
                       activebackground=BG2, relief="flat").pack(anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # Sort by
        tk.Label(ctrl, text="SORT RESULTS BY", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))
        self._sq_sort_var = tk.StringVar(value="combined")
        sort_opts = [
            ("Combined Score",   "combined"),
            ("Gill Score",       "gill"),
            ("Chamath Score",    "chamath"),
            ("Short Interest %", "si"),
            ("Days to Cover",    "dtc"),
        ]
        for label, val in sort_opts:
            tk.Radiobutton(ctrl, text=label, variable=self._sq_sort_var, value=val,
                           font=FONT_SM, bg=BG2, fg=FG, selectcolor=BG3,
                           activebackground=BG2, relief="flat").pack(anchor="w", padx=14)

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # Run button
        self._sq_run_btn = tk.Button(ctrl, text="🎯  Start Squeeze Search",
                                      font=("Consolas",11,"bold"),
                                      bg=ACCENT, fg="#000000",
                                      relief="flat", cursor="hand2",
                                      padx=14, pady=6,
                                      command=self._sq_toggle)
        self._sq_run_btn.pack(fill="x", padx=10, pady=4)

        self._sq_status = tk.Label(ctrl, text="Ready — will scan S&P500",
                                    font=FONT_SM, bg=BG2, fg=FG_DIM, wraplength=230)
        self._sq_status.pack(padx=14, pady=4, anchor="w")

        # ── RIGHT: Results ──────────────────────────────────────────────
        right = tk.Frame(parent, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Progress bar at top
        prog_frame = tk.Frame(right, bg=BG2, pady=6)
        prog_frame.pack(fill="x")

        self._sq_prog_lbl = tk.Label(prog_frame, text="",
                                      font=FONT_SM, bg=BG2, fg=FG_DIM)
        self._sq_prog_lbl.pack(side="left", padx=12)

        self._sq_prog_bar_frame = tk.Frame(prog_frame, bg=BG3, height=6)
        self._sq_prog_bar_frame.pack(side="left", fill="x", expand=True, padx=8, pady=4)
        self._sq_prog_fill = tk.Frame(self._sq_prog_bar_frame, bg=ACCENT, height=6, width=0)
        self._sq_prog_fill.place(x=0, y=0, relheight=1.0)

        tk.Frame(right, bg=BORDER, height=1).pack(fill="x")

        # Results log
        self._sq_log = scrolledtext.ScrolledText(
            right, wrap="word", font=FONT_SM, bg=BG, fg=FG,
            insertbackground=FG, relief="flat", borderwidth=0,
            state="disabled", padx=16, pady=10,
        )
        self._sq_log.pack(fill="both", expand=True)

        # Tags
        for tag, cfg in [
            ("header",   {"font": FONT_LG,                    "foreground": ACCENT}),
            ("dim",      {                                      "foreground": FG_DIM}),
            ("green",    {                                      "foreground": GREEN}),
            ("red",      {                                      "foreground": RED}),
            ("yellow",   {                                      "foreground": YELLOW}),
            ("blue",     {                                      "foreground": BLUE}),
            ("strong",   {"font": ("Consolas",10,"bold"),      "foreground": GREEN}),
            ("watch",    {"font": ("Consolas",10,"bold"),      "foreground": YELLOW}),
            ("pass_tag", {"font": ("Consolas",9),              "foreground": FG_DIM}),
        ]:
            self._sq_log.tag_config(tag, **cfg)

    # ── SQUEEZE SEARCH HELPERS ───────────────────────────────────────────

    def _sq_write(self, text, tag=None):
        self._sq_log.config(state="normal")
        self._sq_log.insert("end", text, tag) if tag else self._sq_log.insert("end", text)
        self._sq_log.see("end")
        self._sq_log.config(state="disabled")
        self.root.update_idletasks()

    def _sq_toggle(self):
        if self._sq_running:
            self._sq_stop = True
            self._sq_run_btn.config(text="⏹ Stopping...", bg=RED, state="disabled")
        else:
            self._sq_start()

    def _sq_start(self):
        try:
            max_stocks  = int(self._sq_max_stocks.get())
            top_n       = int(self._sq_top_results.get())
            min_si      = float(self._sq_min_si.get()) / 100.0
            min_dtc     = float(self._sq_min_dtc.get())
            min_score   = float(self._sq_min_score.get())
        except ValueError:
            self._sq_write("❌ Invalid inputs.\n", "red")
            return

        use_gill    = self._sq_use_gill.get()
        use_chamath = self._sq_use_chamath.get()
        sort_by     = self._sq_sort_var.get()

        if not use_gill and not use_chamath:
            self._sq_write("❌ Select at least one squeeze agent.\n", "red")
            return

        self._sq_running = True
        self._sq_stop    = False
        self._sq_results = []
        self._sq_run_btn.config(bg=RED, fg="#000000", text="⏹ Stop")

        # Clear log
        self._sq_log.config(state="normal")
        self._sq_log.delete("1.0", "end")
        self._sq_log.config(state="disabled")

        threading.Thread(
            target=self._sq_thread,
            args=(max_stocks, top_n, min_si, min_dtc, min_score,
                  use_gill, use_chamath, sort_by),
            daemon=True,
        ).start()

    def _sq_thread(self, max_stocks, top_n, min_si, min_dtc, min_score,
                   use_gill, use_chamath, sort_by):
        try:
            from squeeze_analyzers import (
                run_gill_analysis, run_chamath_analysis,
                fetch_squeeze_metrics
            )

            self._sq_write("\n")
            self._sq_write("  🎯 SQUEEZE SEARCHER\n", "header")
            self._sq_write(f"  Scanning S&P500 — smallest market cap first\n", "dim")
            self._sq_write(f"  Max stocks: {max_stocks}  |  "
                           f"Min SI: {min_si:.0%}  |  Min DTC: {min_dtc:.1f}d  |  "
                           f"Min score: {min_score:.0f}\n\n", "dim")

            # ── Fetch S&P500, sorted smallest market cap first ──
            self._sq_status.config(text="Fetching S&P500 constituent list...")
            sp500_tickers = self._sq_get_sp500_by_marketcap(ascending=True)

            if not sp500_tickers:
                self._sq_write("❌ Could not fetch S&P500 list. Check internet connection.\n", "red")
                return

            sp500_tickers = sp500_tickers[:max_stocks]
            total = len(sp500_tickers)
            self._sq_write(f"  Found {total} tickers to scan\n\n", "dim")

            candidates = []

            for i, ticker in enumerate(sp500_tickers):
                if self._sq_stop:
                    self._sq_write("\n  ⏹ Stopped.\n", "dim")
                    break

                # Update progress bar
                pct = (i + 1) / total
                self._sq_status.config(text=f"[{i+1}/{total}] {ticker}")
                self._sq_prog_lbl.config(text=f"{i+1}/{total}")
                try:
                    bar_w = self._sq_prog_bar_frame.winfo_width()
                    self._sq_prog_fill.place(x=0, y=0,
                                              width=int(bar_w * pct),
                                              relheight=1.0)
                except Exception:
                    pass

                # Quick pre-filter: fetch raw metrics first
                try:
                    metrics = fetch_squeeze_metrics(ticker)
                except Exception as e:
                    self._sq_write(f"  ⚠️  {ticker}: fetch error — {e}\n", "dim")
                    continue

                # Pre-filter on minimum thresholds to skip obvious non-candidates fast
                si = metrics.short_interest_pct or 0
                dtc = metrics.days_to_cover or 0
                if si < min_si or dtc < min_dtc:
                    self._sq_write(
                        f"  ○ {ticker:<6}  SI:{si:.0%}  DTC:{dtc:.1f}d  — below threshold\n",
                        "pass_tag"
                    )
                    continue

                # Run full squeeze analyses
                gill_result    = None
                chamath_result = None
                gill_score     = 0
                chamath_score  = 0

                if use_gill:
                    try:
                        gill_result = run_gill_analysis(ticker)
                        gill_score  = gill_result.total_score
                    except Exception:
                        pass

                if use_chamath:
                    try:
                        chamath_result = run_chamath_analysis(ticker)
                        chamath_score  = chamath_result.total_score
                    except Exception:
                        pass

                combined = (gill_score + chamath_score) / (
                    (1 if use_gill else 0) + (1 if use_chamath else 0)
                )

                if combined < min_score:
                    self._sq_write(
                        f"  ○ {ticker:<6}  Score:{combined:.0f}  SI:{si:.0%}  DTC:{dtc:.1f}d  — score too low\n",
                        "pass_tag"
                    )
                    continue

                # Passed all filters — this is a candidate
                candidates.append({
                    "ticker":   ticker,
                    "company":  metrics.company_name,
                    "sector":   metrics.sector,
                    "gill":     gill_score,
                    "chamath":  chamath_score,
                    "combined": combined,
                    "si":       si,
                    "dtc":      dtc,
                    "ctb":      metrics.ctb_proxy,
                    "price":    metrics.current_price,
                    "mktcap":   metrics.market_cap,
                    "gill_obj": gill_result,
                    "ch_obj":   chamath_result,
                    "verdict":  gill_result.verdict if gill_result else (chamath_result.verdict if chamath_result else ""),
                })

                verdict_tag = "strong" if combined >= 60 else "watch"
                self._sq_write(
                    f"  ✅ {ticker:<6}  Combined:{combined:.0f}  "
                    f"Gill:{gill_score:.0f}  Chamath:{chamath_score:.0f}  "
                    f"SI:{si:.0%}  DTC:{dtc:.1f}d\n",
                    verdict_tag
                )

            # ── Sort and display final results ──
            sort_key = {
                "combined": lambda x: x["combined"],
                "gill":     lambda x: x["gill"],
                "chamath":  lambda x: x["chamath"],
                "si":       lambda x: x["si"],
                "dtc":      lambda x: x["dtc"],
            }.get(sort_by, lambda x: x["combined"])

            candidates.sort(key=sort_key, reverse=True)
            self._sq_results = candidates

            self._sq_write("\n")
            self._sq_write(f"  {'─'*70}\n", "dim")
            self._sq_write(f"  TOP SQUEEZE CANDIDATES\n", "header")
            self._sq_write(f"  Sorted by: {sort_by}  |  "
                           f"Found {len(candidates)} candidates from {i+1} scanned\n\n", "dim")

            if not candidates:
                self._sq_write("  No candidates found matching your criteria.\n", "yellow")
                self._sq_write("  Try lowering Min Short Interest % or Min Score.\n", "dim")
            else:
                # Header
                self._sq_write(
                    f"  {'Rank':<5} {'Ticker':<7} {'Company':<22} {'Sector':<20} "
                    f"{'Comb':>5}  {'Gill':>5}  {'Cham':>5}  "
                    f"{'SI':>6}  {'DTC':>5}  {'CTB':>6}  {'MktCap':>8}\n", "blue"
                )
                self._sq_write(f"  {'─'*105}\n", "dim")

                for rank, c in enumerate(candidates[:top_n], 1):
                    mc_str  = f"${c['mktcap']/1e9:.1f}B" if c.get("mktcap") else "N/A"
                    ctb_str = f"{c['ctb']:.0f}%" if c.get("ctb") else "N/A"
                    tag = "strong" if c["combined"] >= 60 else "watch"

                    self._sq_write(
                        f"  #{rank:<4} {c['ticker']:<7} {c['company'][:22]:<22} "
                        f"{c['sector'][:20]:<20} "
                        f"{c['combined']:>5.0f}  {c['gill']:>5.0f}  {c['chamath']:>5.0f}  "
                        f"{c['si']:>6.1%}  {c['dtc']:>5.1f}  {ctb_str:>6}  {mc_str:>8}\n",
                        tag
                    )

                # Detailed breakdown for top 3
                self._sq_write(f"\n  {'─'*70}\n", "dim")
                self._sq_write(f"  DETAILED BREAKDOWN — TOP 3\n\n", "header")

                from squeeze_analyzers import format_gill_display, format_chamath_display
                for c in candidates[:3]:
                    self._sq_write(f"  {'='*68}\n", "dim")
                    self._sq_write(f"  #{candidates.index(c)+1}  {c['ticker']} — {c['company']}\n", "strong")
                    self._sq_write(f"  Combined Score: {c['combined']:.0f}  |  "
                                   f"Verdict: {c['verdict']}\n\n", "watch")
                    if c.get("gill_obj") and use_gill:
                        self._sq_write(f"  🎮 KEITH GILL\n", "blue")
                        self._sq_write(format_gill_display(c["gill_obj"]), "dim")
                    if c.get("ch_obj") and use_chamath:
                        self._sq_write(f"  💰 CHAMATH\n", "blue")
                        self._sq_write(format_chamath_display(c["ch_obj"]), "dim")

            self._sq_prog_lbl.config(text=f"Done — {len(candidates)} candidates")
            self._sq_status.config(text=f"Done — {len(candidates)} squeeze candidates found")

        except Exception as e:
            import traceback
            self._sq_write(f"\n❌ Error: {e}\n", "red")
            self._sq_write(traceback.format_exc(), "dim")
            self._sq_status.config(text="Error")
        finally:
            self._sq_running = False
            self._sq_stop    = False
            self._sq_run_btn.config(state="normal", text="🎯  Start Squeeze Search",
                                     bg=ACCENT, fg="#000000")

    def _sq_get_sp500_by_marketcap(self, ascending=True) -> list:
        """
        Fetch S&P500 tickers sorted by market cap.
        ascending=True → smallest first (more squeeze candidates at small cap).
        Returns list of ticker strings.
        """
        import yfinance as yf

        tickers = []

        # Method 1: Wikipedia via requests + html.parser (no lxml needed)
        try:
            import pandas as pd
            import requests
            resp = requests.get(
                "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
                headers={"User-Agent": "Mozilla/5.0"}, timeout=15
            )
            sp500 = pd.read_html(resp.text, attrs={"id": "constituents"})[0]
            sym_col = "Symbol" if "Symbol" in sp500.columns else sp500.columns[0]
            tickers = sp500[sym_col].str.replace(".", "-", regex=False).tolist()
            self._sq_write(f"  📋 Wikipedia: {len(tickers)} S&P500 tickers\n", "dim")
        except Exception as e:
            self._sq_write(f"  ⚠️  Wikipedia method failed: {e}\n", "dim")

        # Method 2: yfinance screener for US large/mid cap equities
        if not tickers:
            try:
                import yfinance as yf
                screener = yf.Screener()
                screener.set_body({
                    "offset": 0, "size": 500,
                    "sortField": "intradaymarketcap", "sortType": "DESC",
                    "quoteType": "EQUITY",
                    "query": {
                        "operator": "and",
                        "operands": [
                            {"operator": "eq",  "operands": ["region", "us"]},
                            {"operator": "gt",  "operands": ["intradaymarketcap", 2_000_000_000]},
                        ]
                    },
                    "userId": "", "userIdType": "guest"
                })
                quotes = screener.response.get("quotes", [])
                tickers = [q["symbol"] for q in quotes if q.get("symbol")]
                self._sq_write(f"  📋 Screener: {len(tickers)} large/mid cap tickers\n", "dim")
            except Exception as e:
                self._sq_write(f"  ⚠️  Screener method failed: {e}\n", "dim")

        # Method 3: Hardcoded S&P500-representative list as last resort
        if not tickers:
            self._sq_write("  ⚠️  Using built-in representative ticker list\n", "yellow")
            tickers = [
                # Small/micro cap (most squeeze candidates)
                "BBAI","MARA","RIOT","CLSK","CIFR","HUT","BTBT","ARBK",
                "SRRK","NKLA","RIDE","WKHS","GOEV","MULN","FFIE","CENN",
                "SPCE","JOBY","LILM","ACHR","EVTL","EHGO","GREE","HIVE",
                # Mid cap
                "GME","AMC","BBBY","KOSS","EXPR","NAKD","SNDL","CLOV",
                "WISH","WOOF","PAYO","OPEN","OFSG","PRCH","SKLZ","DKNG",
                # Large cap (lower squeeze probability but included)
                "NVDA","AMD","TSLA","META","GOOGL","MSFT","AAPL","AMZN",
                "JPM","BAC","WFC","GS","MS","C","USB","PNC","TFC","FITB",
                "XOM","CVX","COP","SLB","HAL","DVN","EOG","PXD","MPC","VLO",
                "LLY","UNH","JNJ","ABBV","MRK","ABT","TMO","DHR","ISRG","REGN",
            ]

        if not tickers:
            self._sq_write("  ❌ Could not build ticker list from any source.\n", "red")
            return []

        self._sq_write(f"  📋 Got {len(tickers)} S&P500 tickers — fetching market caps...\n", "dim")

        # Fetch market caps in bulk using yfinance download
        # Use info for a batch — this is slow but necessary for sorting
        # Speed optimization: use fast_info which is cached
        ticker_caps = {}
        batch_size = 50

        for i in range(0, min(len(tickers), 500), batch_size):
            batch = tickers[i:i+batch_size]
            if self._sq_stop:
                break
            try:
                # Download to get tickers validated, then pull fast_info
                for t in batch:
                    try:
                        fast = yf.Ticker(t).fast_info
                        mc = getattr(fast, "market_cap", None)
                        ticker_caps[t] = mc if mc else 0
                    except Exception:
                        ticker_caps[t] = 0
            except Exception:
                for t in batch:
                    ticker_caps[t] = 0

            self._sq_status.config(text=f"Sorting by market cap... {i+batch_size}/{len(tickers)}")

        # Sort by market cap
        sorted_tickers = sorted(
            ticker_caps.keys(),
            key=lambda t: ticker_caps.get(t, 0),
            reverse=not ascending
        )

        cap_order = "smallest → largest" if ascending else "largest → smallest"
        self._sq_write(f"  ✅ Sorted {len(sorted_tickers)} tickers by market cap ({cap_order})\n\n", "dim")
        return sorted_tickers

    def _build_squeeze_analyzer_tab(self):
        parent = self.tab_squeeze_single
        self._sa_running = False
        self._sa_stop    = False
        self._sa_results = None
        main = tk.Frame(parent, bg=BG)
        main.pack(fill="both", expand=True)
        left = tk.Frame(main, bg=BG)
        left.pack(side="left", fill="both", expand=True)
        self._sa_chat = scrolledtext.ScrolledText(
            left, wrap="word", font=FONT, bg=BG, fg=FG,
            insertbackground=FG, selectbackground=BORDER,
            relief="flat", borderwidth=0, state="disabled", padx=20, pady=16)
        self._sa_chat.pack(fill="both", expand=True)
        for tag, cfg in [
            ("header",       {"font": FONT_LG, "foreground": ACCENT}),
            ("dim",          {"foreground": FG_DIM}),
            ("green",        {"foreground": GREEN}),
            ("red",          {"foreground": RED}),
            ("yellow",       {"foreground": YELLOW}),
            ("blue",         {"foreground": BLUE}),
            ("score_strong", {"font": ("Consolas",12,"bold"), "foreground": GREEN}),
            ("score_watch",  {"font": ("Consolas",11,"bold"), "foreground": YELLOW}),
            ("score_pass",   {"font": ("Consolas",11,"bold"), "foreground": RED}),
            ("claude",       {"font": ("Consolas",10), "foreground": TEAL}),
        ]:
            self._sa_chat.tag_config(tag, **cfg)
        sb = tk.Frame(main, bg=BG2, width=250)
        sb.pack(side="right", fill="y")
        sb.pack_propagate(False)
        tk.Label(sb, text="SQUEEZE SCORE", font=FONT_SM, bg=BG2, fg=FG_DIM).pack(pady=(14,2), padx=10, anchor="w")
        box = tk.Frame(sb, bg=BG3)
        box.pack(fill="x", padx=6, pady=2)
        self._sa_lbl_ticker  = tk.Label(box, text="—", font=FONT_LG, bg=BG3, fg=ACCENT)
        self._sa_lbl_ticker.pack(pady=(8,0))
        self._sa_lbl_gill    = tk.Label(box, text="Gill: —", font=("Consolas",10,"bold"), bg=BG3, fg=FG_DIM)
        self._sa_lbl_gill.pack()
        self._sa_lbl_chamath = tk.Label(box, text="Chamath: —", font=("Consolas",10,"bold"), bg=BG3, fg=FG_DIM)
        self._sa_lbl_chamath.pack()
        self._sa_lbl_combined = tk.Label(box, text="—", font=("Consolas",28,"bold"), bg=BG3, fg=FG_DIM)
        self._sa_lbl_combined.pack(pady=(4,0))
        self._sa_lbl_verdict = tk.Label(box, text="—", font=("Consolas",9,"bold"), bg=BG3, fg=FG_DIM, wraplength=220)
        self._sa_lbl_verdict.pack(pady=(0,8), padx=6)
        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=6, pady=6)
        tk.Label(sb, text="KEY METRICS", font=FONT_SM, bg=BG2, fg=FG_DIM).pack(anchor="w", padx=10, pady=(0,4))
        self._sa_metrics_frame = tk.Frame(sb, bg=BG2)
        self._sa_metrics_frame.pack(fill="x", padx=6)
        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=6, pady=6)
        tk.Label(sb, text="AGENTS", font=FONT_SM, bg=BG2, fg=FG_DIM).pack(anchor="w", padx=10, pady=(0,4))
        self._sa_use_gill    = tk.BooleanVar(value=True)
        self._sa_use_chamath = tk.BooleanVar(value=True)
        af2 = tk.Frame(sb, bg=BG2)
        af2.pack(fill="x", padx=8)
        for text, var in [("🎮 Keith Gill (DFV)", self._sa_use_gill),
                           ("💰 Chamath Palihapitiya", self._sa_use_chamath)]:
            tk.Checkbutton(af2, text=text, variable=var, font=FONT_SM,
                           bg=BG2, fg=FG, selectcolor=BG3,
                           activebackground=BG2, relief="flat").pack(anchor="w")
        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=6, pady=6)
        tk.Button(sb, text="Clear", font=FONT_SM, bg=BG3, fg=FG_DIM,
                  relief="flat", cursor="hand2", command=self._sa_clear).pack(fill="x", padx=6, pady=2)
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x")
        bot = tk.Frame(parent, bg=BG2, pady=8)
        bot.pack(fill="x")
        r1 = tk.Frame(bot, bg=BG2)
        r1.pack(fill="x", padx=12, pady=(0,4))
        tk.Label(r1, text="Ticker:", font=FONT, bg=BG2, fg=FG_DIM).pack(side="left")
        self._sa_ticker_var = tk.StringVar()
        self._sa_ticker_entry = tk.Entry(r1, textvariable=self._sa_ticker_var, font=FONT,
                                          bg=BG3, fg=FG, insertbackground=FG, relief="flat", bd=6, width=12)
        self._sa_ticker_entry.pack(side="left", padx=6)
        self._sa_ticker_entry.bind("<Return>", lambda e: self._sa_toggle())
        self._sa_run_btn = tk.Button(r1, text="▶  Analyze", font=("Consolas",11,"bold"),
                                      bg=ACCENT, fg="#000000", relief="flat",
                                      cursor="hand2", padx=14, pady=4, command=self._sa_toggle)
        self._sa_run_btn.pack(side="left", padx=4)
        self._sa_status_lbl = tk.Label(r1, text="Ready", font=FONT_SM, bg=BG2, fg=FG_DIM)
        self._sa_status_lbl.pack(side="left", padx=10)
        tk.Frame(bot, bg=BORDER, height=1).pack(fill="x")
        r2 = tk.Frame(bot, bg=BG, pady=6)
        r2.pack(fill="x", padx=12)
        tk.Label(r2, text="Ask Claude:", font=FONT, bg=BG, fg=FG_DIM).pack(side="left")
        self._sa_qa_var = tk.StringVar()
        self._sa_qa_entry = tk.Entry(r2, textvariable=self._sa_qa_var, font=FONT,
                                      bg=BG3, fg=FG_DIM, insertbackground=FG, relief="flat", bd=6, state="disabled")
        self._sa_qa_entry.pack(side="left", fill="x", expand=True, padx=6)
        self._sa_qa_entry.bind("<Return>", lambda e: self._sa_ask_claude())
        self._sa_qa_btn = tk.Button(r2, text="💬 Ask", font=("Consolas",10,"bold"),
                                     bg=BG3, fg=FG_DIM, relief="flat", cursor="hand2",
                                     padx=10, pady=4, state="disabled", command=self._sa_ask_claude)
        self._sa_qa_btn.pack(side="left", padx=4)

    def _sa_w(self, text, tag=None):
        def _do():
            self._sa_chat.config(state="normal")
            if tag: self._sa_chat.insert("end", text, tag)
            else:   self._sa_chat.insert("end", text)
            self._sa_chat.see("end")
            self._sa_chat.config(state="disabled")
        self.root.after(0, _do)

    def _sa_rule(self, label=""):
        pad = max(0, (60 - len(label) - 2) // 2) if label else 0
        self._sa_w(f"\n{'─'*pad} {label} {'─'*pad}\n\n" if label else f"\n{'─'*60}\n\n", "dim")

    def _sa_clear(self):
        self._sa_chat.config(state="normal")
        self._sa_chat.delete("1.0", "end")
        self._sa_chat.config(state="disabled")
        self._sa_results = None
        self._sa_lbl_ticker.config(text="—", fg=ACCENT)
        self._sa_lbl_gill.config(text="Gill: —", fg=FG_DIM)
        self._sa_lbl_chamath.config(text="Chamath: —", fg=FG_DIM)
        self._sa_lbl_combined.config(text="—", fg=FG_DIM)
        self._sa_lbl_verdict.config(text="—", fg=FG_DIM)
        for w in self._sa_metrics_frame.winfo_children(): w.destroy()
        self._sa_qa_entry.config(state="disabled", fg=FG_DIM)
        self._sa_qa_btn.config(state="disabled", bg=BG3, fg=FG_DIM)

    def _sa_toggle(self):
        if self._sa_running:
            self._sa_stop = True
            self._sa_run_btn.config(text="⏹ Stopping...", bg=RED, state="disabled")
        else:
            ticker = self._sa_ticker_var.get().strip().upper()
            if not ticker: self._sa_ticker_entry.focus(); return
            self._sa_start(ticker)

    def _sa_start(self, ticker):
        self._sa_running = True
        self._sa_stop    = False
        self._sa_results = None
        self._sa_run_btn.config(bg=RED, fg="#000000", text="⏹ Stop")
        self._sa_ticker_entry.config(state="disabled")
        self._sa_qa_entry.config(state="disabled", fg=FG_DIM)
        self._sa_qa_btn.config(state="disabled", bg=BG3, fg=FG_DIM)
        self._sa_status_lbl.config(text=f"Analyzing {ticker}...")
        self._sa_chat.config(state="normal")
        self._sa_chat.delete("1.0", "end")
        self._sa_chat.config(state="disabled")
        threading.Thread(target=self._sa_thread, args=(ticker,), daemon=True).start()

    def _sa_thread(self, ticker):
        try:
            from squeeze_analyzers import (run_gill_analysis, run_chamath_analysis,
                                            format_gill_display, format_chamath_display,
                                            fetch_squeeze_metrics)
            use_gill    = self._sa_use_gill.get()
            use_chamath = self._sa_use_chamath.get()
            self._sa_w(f"\n🔬  SQUEEZE ANALYSIS — {ticker}\n", "header")
            self._sa_rule()
            self.root.after(0, lambda: self._sa_status_lbl.config(text="Fetching data..."))
            metrics = fetch_squeeze_metrics(ticker)
            if metrics.fetch_errors:
                self._sa_w(f"  ⚠️  {'; '.join(metrics.fetch_errors[:2])}\n", "yellow")
            self._sa_rule("📊 Short Interest Data")
            self._sa_w(f"  Company:  {metrics.company_name}\n", "blue")
            self._sa_w(f"  Sector:   {metrics.sector}\n", "dim")
            if metrics.current_price: self._sa_w(f"  Price:    ${metrics.current_price:.2f}\n", "dim")
            if metrics.market_cap:    self._sa_w(f"  Mkt Cap:  ${metrics.market_cap/1e9:.2f}B\n", "dim")
            self._sa_w("\n")
            si_note = f" [{metrics.si_data_quality}]" if metrics.si_data_quality else ""
            self._sa_w(f"  {'Metric':<28} {'Value':<16} Threshold\n", "header")
            self._sa_w(f"  {'─'*60}\n", "dim")
            def mrow(lbl, val, ctx, thresh=None, raw=None):
                tag = "green" if (thresh and raw and raw >= thresh) else ("yellow" if thresh else "dim")
                self._sa_w(f"  {lbl:<28} {val:<16} {ctx}\n", tag)
            mrow("Short Interest % Float", f"{metrics.short_interest_pct:.1%}{si_note}" if metrics.short_interest_pct else "N/A", "> 20% (Gill)", 0.20, metrics.short_interest_pct)
            mrow("Days to Cover (DTC)", f"{metrics.days_to_cover:.1f}d" if metrics.days_to_cover else "N/A", "> 5 days", 5.0, metrics.days_to_cover)
            mrow("CTB Proxy", f"{metrics.ctb_proxy:.1f}%" if metrics.ctb_proxy else "N/A", "> 10%", 10.0, metrics.ctb_proxy)
            mrow("Shares Short", f"{metrics.shares_short:,}" if metrics.shares_short else "N/A", "")
            mrow("Float Shares", f"{metrics.float_shares:,}" if metrics.float_shares else "N/A", "")
            mrow("Short Change", f"{metrics.short_change_pct:+.1%}" if metrics.short_change_pct is not None else "N/A", "+ = adding")
            mrow("FTD % Float", f"{metrics.ftd_pct_float:.3%}" if metrics.ftd_pct_float else "N/A", "SEC data")
            mrow("Volume Surge", f"{metrics.volume_surge:.2f}x" if metrics.volume_surge else "N/A", "vs 20d avg")
            mrow("RSI (14)", f"{metrics.rsi_14:.0f}" if metrics.rsi_14 else "N/A", "< 70 preferred")
            mrow("1-Month Return", f"{metrics.price_change_1m:+.1%}" if metrics.price_change_1m is not None else "N/A", "")
            mrow("3-Month Return", f"{metrics.price_change_3m:+.1%}" if metrics.price_change_3m is not None else "N/A", "")
            self._sa_w("\n")
            if self._sa_stop: self._sa_w("  ⏹ Stopped.\n", "dim"); return
            gill_score = chamath_score = 0.0
            gill_result = chamath_result = None
            if use_gill:
                self.root.after(0, lambda: self._sa_status_lbl.config(text="Running Gill..."))
                self._sa_rule("🎮 Keith Gill — DeepFuckingValue")
                gill_result = run_gill_analysis(ticker)
                gill_score  = gill_result.total_score
                self._sa_w(format_gill_display(gill_result), "dim")
            if not self._sa_stop and use_chamath:
                self.root.after(0, lambda: self._sa_status_lbl.config(text="Running Chamath..."))
                self._sa_rule("💰 Chamath Palihapitiya")
                chamath_result = run_chamath_analysis(ticker)
                chamath_score  = chamath_result.total_score
                self._sa_w(format_chamath_display(chamath_result), "dim")
            active   = sum([use_gill, use_chamath])
            combined = (gill_score + chamath_score) / max(active, 1)
            self._sa_rule("📋 Combined Verdict")
            self._sa_w(f"  {'Agent':<22} {'Score':>8}  Verdict\n", "blue")
            self._sa_w(f"  {'─'*55}\n", "dim")
            if use_gill and gill_result:
                t = "green" if gill_score >= 55 else ("yellow" if gill_score >= 38 else "dim")
                self._sa_w(f"  {'Keith Gill':<22} {gill_score:>7.0f}  {gill_result.verdict} ({gill_result.conviction})\n", t)
            if use_chamath and chamath_result:
                t = "green" if chamath_score >= 70 else ("yellow" if chamath_score >= 50 else "dim")
                self._sa_w(f"  {'Chamath':<22} {chamath_score:>7.0f}  {chamath_result.verdict}\n", t)
            self._sa_w(f"  {'─'*55}\n", "dim")
            comb_tag = "score_strong" if combined >= 65 else ("score_watch" if combined >= 45 else "score_pass")
            self._sa_w(f"  {'COMBINED':<22} {combined:>7.0f}\n", comb_tag)
            overall = ("SQUEEZE CANDIDATE 🔥" if combined >= 65 else
                       ("WATCH — Setup Building ⚠️" if combined >= 45 else "PASS — No Squeeze Setup"))
            self._sa_w(f"\n  {overall}\n\n", comb_tag)
            self._sa_rule()
            self._sa_results = {"ticker": ticker, "metrics": metrics,
                                 "gill": gill_result, "chamath": chamath_result,
                                 "combined": combined, "overall": overall}
            def _update_sb():
                cc = GREEN if combined >= 65 else (YELLOW if combined >= 45 else RED)
                self._sa_lbl_ticker.config(text=ticker)
                self._sa_lbl_combined.config(text=f"{combined:.0f}", fg=cc)
                self._sa_lbl_verdict.config(text=overall, fg=cc)
                if use_gill:
                    gc = GREEN if gill_score >= 55 else (YELLOW if gill_score >= 38 else RED)
                    self._sa_lbl_gill.config(text=f"Gill:    {gill_score:.0f}/100", fg=gc)
                if use_chamath:
                    cc2 = GREEN if chamath_score >= 70 else (YELLOW if chamath_score >= 50 else RED)
                    self._sa_lbl_chamath.config(text=f"Chamath: {chamath_score:.0f}/100", fg=cc2)
                for w in self._sa_metrics_frame.winfo_children(): w.destroy()
                for lbl, val, col in [
                    ("SI% Float", f"{metrics.short_interest_pct:.1%}" if metrics.short_interest_pct else "N/A", GREEN if (metrics.short_interest_pct or 0) >= 0.20 else YELLOW),
                    ("DTC",       f"{metrics.days_to_cover:.1f}d" if metrics.days_to_cover else "N/A", GREEN if (metrics.days_to_cover or 0) >= 5 else YELLOW),
                    ("CTB",       f"{metrics.ctb_proxy:.0f}%" if metrics.ctb_proxy else "N/A", GREEN if (metrics.ctb_proxy or 0) >= 10 else YELLOW),
                    ("FTD",       f"{metrics.ftd_pct_float:.3%}" if metrics.ftd_pct_float else "None", FG_DIM),
                    ("RSI",       f"{metrics.rsi_14:.0f}" if metrics.rsi_14 else "N/A", FG_DIM),
                    ("Vol Surge", f"{metrics.volume_surge:.1f}x" if metrics.volume_surge else "N/A", GREEN if (metrics.volume_surge or 0) >= 2.0 else FG_DIM),
                ]:
                    rw = tk.Frame(self._sa_metrics_frame, bg=BG2)
                    rw.pack(fill="x", pady=1)
                    tk.Label(rw, text=lbl, font=FONT_SM, bg=BG2, fg=FG_DIM, width=10, anchor="w").pack(side="left")
                    tk.Label(rw, text=val, font=FONT_SM, bg=BG2, fg=col, anchor="e").pack(side="right")
                self._sa_qa_entry.config(state="normal", fg=FG)
                self._sa_qa_btn.config(state="normal", bg="#238636", fg="#FFFFFF")
            self.root.after(0, _update_sb)
            self._sa_w("  💬 Ask Claude about this analysis below\n", "dim")
            self.root.after(0, lambda: self._sa_status_lbl.config(text=f"Done — {combined:.0f}  {overall}"))
        except Exception as e:
            import traceback
            self._sa_w(f"\n❌ Error: {e}\n", "red")
            self._sa_w(traceback.format_exc() + "\n", "dim")
            self.root.after(0, lambda: self._sa_status_lbl.config(text="Error"))
        finally:
            self._sa_running = False
            self._sa_stop    = False
            self.root.after(0, lambda: [
                self._sa_run_btn.config(state="normal", text="▶  Analyze", bg=ACCENT, fg="#000000"),
                self._sa_ticker_entry.config(state="normal"),
                self._sa_ticker_var.set(""),
                self._sa_ticker_entry.focus(),
            ])

    def _sa_ask_claude(self):
        if self._sa_running or not self._sa_results: return
        question = self._sa_qa_var.get().strip()
        if not question: self._sa_qa_entry.focus(); return
        self._sa_running = True
        self._sa_qa_btn.config(state="disabled", text="⏳")
        self._sa_qa_entry.config(state="disabled")
        self._sa_run_btn.config(state="disabled")
        threading.Thread(target=self._sa_claude_thread, args=(question,), daemon=True).start()

    def _sa_claude_thread(self, question):
        try:
            r = self._sa_results
            m = r["metrics"]
            lines = [f"SQUEEZE ANALYSIS — {r['ticker']}", f"Combined: {r['combined']:.0f}/100 — {r['overall']}",
                     f"SI: {m.short_interest_pct:.1%}" if m.short_interest_pct else "SI: N/A",
                     f"DTC: {m.days_to_cover:.1f}d" if m.days_to_cover else "DTC: N/A"]
            if r.get("gill"):
                lines += [f"GILL: {r['gill'].total_score:.0f}/100 — {r['gill'].verdict}",
                          "Green: " + "; ".join(r['gill'].green_flags[:3])]
            if r.get("chamath"):
                lines.append(f"CHAMATH: {r['chamath'].total_score:.0f}/100 — {r['chamath'].verdict}")
            context = "\n".join(l for l in lines if l)
            self._sa_rule("Claude Q&A")
            self._sa_w(f"  Q: {question}\n\n", "blue")
            self._sa_w("  ⏳ Thinking...\n", "dim")
            answer = ask_lm_studio(question, context, self.portfolio_ctx)
            self._sa_chat.config(state="normal")
            pos = self._sa_chat.search("  ⏳ Thinking...", "1.0", "end")
            if pos: self._sa_chat.delete(pos, f"{pos} lineend+1c")
            self._sa_chat.config(state="disabled")
            self._sa_w(f"  {answer}\n\n", "claude")
        except Exception as e:
            self._sa_w(f"  ❌ Error: {e}\n", "red")
        finally:
            self._sa_running = False
            self.root.after(0, lambda: [
                self._sa_qa_btn.config(state="normal", text="💬 Ask", bg="#238636", fg="#FFFFFF"),
                self._sa_qa_entry.config(state="normal", fg=FG),
                self._sa_run_btn.config(state="normal"),
                self._sa_qa_var.set(""),
                self._sa_qa_entry.focus(),
            ])

    def _init_prices(self):
        self.port_label.config(text="📡 Fetching prices...", fg=FG_DIM)
        if os.path.exists(PORTFOLIO_FILE):
            fetch_live_prices(PORTFOLIO_FILE)
            self.portfolio_ctx = load_portfolio_context(PORTFOLIO_FILE)
            self.port_label.config(text="📋 Portfolio loaded", fg=GREEN)
        else:
            self.port_label.config(text="📋 No portfolio.xlsx", fg=YELLOW)

    # ── BUILD UI ──────────────────────────────
    def _build_ui(self):
        import tkinter.ttk as ttk

        # ── SHARED TOP BAR ──
        top = tk.Frame(self.root, bg=BG2, pady=8)
        top.pack(fill="x")
        tk.Label(top, text="💼 INVESTOR ROUNDTABLE", font=FONT_HD,
                 bg=BG2, fg=ACCENT).pack(side="left", padx=16)

        # Portfolio reload (right side)
        self.port_label = tk.Label(top, text="📋 Loading...", font=FONT_SM, bg=BG2, fg=FG_DIM)
        self.port_label.pack(side="right", padx=8)
        tk.Button(top, text="Reload Portfolio", font=FONT_SM,
                  bg=BG3, fg=FG, relief="flat", cursor="hand2",
                  command=self._reload_portfolio).pack(side="right", padx=4)

        # Backend / model switcher
        tk.Frame(top, bg=BORDER, width=1).pack(side="right", fill="y", padx=6, pady=4)

        self.conn_lbl = tk.Label(top, text="⏳ checking...", font=FONT_SM, bg=BG2, fg=FG_DIM)
        self.conn_lbl.pack(side="right", padx=4)

        self._model_options = {}
        self._model_options["🖥  Local (LM Studio)"] = ("local", "local-model")
        try:
            for name, mid in GROQ_MODELS.items():
                self._model_options[f"⚡ Groq — {name}"] = ("groq", mid)
            for name, mid in TOGETHER_MODELS.items():
                self._model_options[f"☁  Together — {name}"] = ("together", mid)
        except NameError:
            pass   # GROQ_MODELS / TOGETHER_MODELS may not be defined in this version

        self._backend_var = tk.StringVar(value="🖥  Local (LM Studio)")
        backend_menu = tk.OptionMenu(
            top, self._backend_var,
            *self._model_options.keys(),
            command=self._on_backend_change if hasattr(self, "_on_backend_change") else lambda x: None,
        )
        backend_menu.config(
            font=FONT_SM, bg=BG3, fg=FG, activebackground=BG2,
            activeforeground=ACCENT, relief="flat",
            highlightthickness=0, bd=0,
        )
        backend_menu["menu"].config(bg=BG3, fg=FG, font=FONT_SM, activebackground=BORDER)
        backend_menu.pack(side="right", padx=2)
        tk.Label(top, text="Model:", font=FONT_SM, bg=BG2, fg=FG_DIM).pack(side="right", padx=(8,2))

        self._apikey_frame = tk.Frame(top, bg=BG2)
        self._apikey_frame.pack(side="right", padx=4)
        self._apikey_var = tk.StringVar()
        self._apikey_entry = tk.Entry(self._apikey_frame, textvariable=self._apikey_var,
                                       font=FONT_SM, bg=BG3, fg=FG, show="*",
                                       relief="flat", bd=4, width=24)

        # ── NOTEBOOK TABS ──
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TNotebook",         background=BG2, borderwidth=0)
        style.configure("TNotebook.Tab",     background=BG3, foreground=FG_DIM,
                         padding=[16, 6], font=FONT_SM)
        style.map("TNotebook.Tab",
                  background=[("selected", BG),  ("active", BG3)],
                  foreground=[("selected", ACCENT), ("active", FG)])

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, pady=0)

        # Tab 1: Stock Analysis
        self.tab_analysis = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(self.tab_analysis, text="  📊 Stock Analysis  ")

        # Tab 2: Portfolio Builder
        self.tab_portfolio = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(self.tab_portfolio, text="  🏗  Portfolio Builder  ")

        # Tab 3: Squeeze Searcher
        self.tab_squeeze = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(self.tab_squeeze, text="  🎯  Squeeze Searcher  ")

        self.tab_squeeze_single = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(self.tab_squeeze_single, text="  🔬  Squeeze Analyzer  ")

        # Build each tab
        self._build_analysis_tab()
        self._build_portfolio_tab()
        self._build_squeeze_tab()
        self._build_squeeze_analyzer_tab()

    def _build_analysis_tab(self):
        """Stock Analysis tab — composite score dashboard."""
        threading.Thread(target=self._check_backend, daemon=True).start()

        # MAIN LAYOUT — inside Stock Analysis tab
        main = tk.Frame(self.tab_analysis, bg=BG)
        main.pack(fill="both", expand=True)

        # LEFT: output
        left = tk.Frame(main, bg=BG)
        left.pack(side="left", fill="both", expand=True)

        self.chat = scrolledtext.ScrolledText(
            left, wrap="word", font=FONT, bg=BG, fg=FG,
            insertbackground=FG, selectbackground=BORDER,
            relief="flat", borderwidth=0, state="disabled",
            padx=20, pady=16,
        )
        self.chat.pack(fill="both", expand=True)

        # Text tags
        for tag, cfg in [
            ("header",       {"font": FONT_LG,                     "foreground": ACCENT}),
            ("dim",          {                                       "foreground": FG_DIM}),
            ("green",        {                                       "foreground": GREEN}),
            ("red",          {                                       "foreground": RED}),
            ("yellow",       {                                       "foreground": YELLOW}),
            ("blue",         {                                       "foreground": BLUE}),
            ("teal",         {                                       "foreground": TEAL}),
            ("score_strong", {"font": ("Consolas",12,"bold"),       "foreground": GREEN}),
            ("score_buy",    {"font": ("Consolas",11,"bold"),       "foreground": GREEN}),
            ("score_watch",  {"font": ("Consolas",11,"bold"),       "foreground": YELLOW}),
            ("score_avoid",  {"font": ("Consolas",11,"bold"),       "foreground": RED}),
            ("claude",       {"font": ("Consolas",10),              "foreground": TEAL}),
        ]:
            self.chat.tag_config(tag, **cfg)

        # RIGHT: sidebar
        sb = tk.Frame(main, bg=BG2, width=250)
        sb.pack(side="right", fill="y")
        sb.pack_propagate(False)

        tk.Label(sb, text="LAST SCORE", font=FONT_SM, bg=BG2, fg=FG_DIM).pack(pady=(14,2), padx=10, anchor="w")

        self._score_box = tk.Frame(sb, bg=BG3)
        self._score_box.pack(fill="x", padx=6, pady=2)

        self._lbl_ticker = tk.Label(self._score_box, text="—", font=FONT_LG, bg=BG3, fg=ACCENT)
        self._lbl_ticker.pack(pady=(8,0))
        self._lbl_num    = tk.Label(self._score_box, text="—", font=("Consolas",34,"bold"), bg=BG3, fg=FG_DIM)
        self._lbl_num.pack()
        self._lbl_sig    = tk.Label(self._score_box, text="—", font=("Consolas",11,"bold"), bg=BG3, fg=FG_DIM)
        self._lbl_sig.pack()
        self._lbl_fit    = tk.Label(self._score_box, text="", font=FONT_SM, bg=BG3, fg=BLUE, wraplength=210)
        self._lbl_fit.pack(pady=(0,8), padx=6)

        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=6, pady=6)
        tk.Label(sb, text="COMPONENTS", font=FONT_SM, bg=BG2, fg=FG_DIM).pack(anchor="w", padx=10, pady=(0,4))

        self._comp_frame = tk.Frame(sb, bg=BG2)
        self._comp_frame.pack(fill="x", padx=6)

        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=6, pady=6)
        tk.Button(sb, text="Clear", font=FONT_SM, bg=BG3, fg=FG_DIM,
                  relief="flat", cursor="hand2", command=self._clear).pack(fill="x", padx=6, pady=2)

        # BOTTOM INPUTS — inside Stock Analysis tab
        tk.Frame(self.tab_analysis, bg=BORDER, height=1).pack(fill="x")
        bot = tk.Frame(self.tab_analysis, bg=BG2, pady=8)
        bot.pack(fill="x")

        r1 = tk.Frame(bot, bg=BG2)
        r1.pack(fill="x", padx=12, pady=(0,4))

        tk.Label(r1, text="Ticker:", font=FONT, bg=BG2, fg=FG_DIM).pack(side="left")
        self.ticker_var = tk.StringVar()
        self.ticker_entry = tk.Entry(r1, textvariable=self.ticker_var, font=FONT,
                                      bg=BG3, fg=FG, insertbackground=FG,
                                      relief="flat", bd=6, width=12)
        self.ticker_entry.pack(side="left", padx=6)
        self.ticker_entry.bind("<Return>", lambda e: self._toggle_run())

        self.run_btn = tk.Button(r1, text="▶  Analyze", font=("Consolas",11,"bold"),
                                  bg=ACCENT, fg="#000000", relief="flat",
                                  cursor="hand2", padx=14, pady=4,
                                  command=self._toggle_run)
        self.run_btn.pack(side="left", padx=4)

        # ── MODE TOGGLE SWITCH ──
        mode_frame = tk.Frame(r1, bg=BG2)
        mode_frame.pack(side="left", padx=12)

        def _update_mode_display(*_):
            m = self.mode_var.get()
            if m == "composite":
                lbl_composite.config(fg=ACCENT, font=("Consolas",9,"bold"))
                lbl_long.config(fg=FG_DIM, font=FONT_SM)
                toggle_btn.config(bg="#238636")
                self.run_btn.config(text="▶  Score")
            else:
                lbl_composite.config(fg=FG_DIM, font=FONT_SM)
                lbl_long.config(fg=BLUE, font=("Consolas",9,"bold"))
                toggle_btn.config(bg=BLUE)
                self.run_btn.config(text="▶  Deep Dive")

        def _toggle_mode():
            self.mode_var.set("long" if self.mode_var.get() == "composite" else "composite")
            _update_mode_display()

        lbl_composite = tk.Label(mode_frame, text="Score", font=("Consolas",9,"bold"),
                                  bg=BG2, fg=ACCENT, cursor="hand2")
        lbl_composite.pack(side="left", padx=(0,4))
        lbl_composite.bind("<Button-1>", lambda e: [self.mode_var.set("composite"), _update_mode_display()])

        toggle_btn = tk.Button(mode_frame, text="  ", width=3, relief="flat",
                                bg="#238636", cursor="hand2", command=_toggle_mode)
        toggle_btn.pack(side="left")

        lbl_long = tk.Label(mode_frame, text="Deep Dive", font=FONT_SM,
                             bg=BG2, fg=FG_DIM, cursor="hand2")
        lbl_long.pack(side="left", padx=(4,0))
        lbl_long.bind("<Button-1>", lambda e: [self.mode_var.set("long"), _update_mode_display()])

        self.status_lbl = tk.Label(r1, text="Ready", font=FONT_SM, bg=BG2, fg=FG_DIM)
        self.status_lbl.pack(side="left", padx=10)

        tk.Frame(bot, bg=BORDER, height=1).pack(fill="x")
        r2 = tk.Frame(bot, bg=BG, pady=6)
        r2.pack(fill="x", padx=12)

        tk.Label(r2, text="Ask Claude:", font=FONT, bg=BG, fg=FG_DIM).pack(side="left")
        self.qa_var = tk.StringVar()
        self.qa_entry = tk.Entry(r2, textvariable=self.qa_var, font=FONT,
                                  bg=BG3, fg=FG_DIM, insertbackground=FG,
                                  relief="flat", bd=6, state="disabled")
        self.qa_entry.pack(side="left", fill="x", expand=True, padx=6)
        self.qa_entry.bind("<Return>", lambda e: self._ask_claude())

        self.qa_btn = tk.Button(r2, text="💬 Ask", font=("Consolas",10,"bold"),
                                 bg=BG3, fg=FG_DIM, relief="flat",
                                 cursor="hand2", padx=10, pady=4,
                                 state="disabled", command=self._ask_claude)
        self.qa_btn.pack(side="left", padx=4)

    # ── UTILITIES ────────────────────────────
    def _w(self, text, tag=None):
        self.chat.config(state="normal")
        self.chat.insert("end", text, tag) if tag else self.chat.insert("end", text)
        self.chat.see("end")
        self.chat.config(state="disabled")
        self.root.update_idletasks()

    def _rule(self, label=""):
        if label:
            pad = max(0, (62 - len(label) - 2) // 2)
            self._w(f"\n{'─'*pad} {label} {'─'*pad}\n\n", "dim")
        else:
            self._w(f"\n{'─'*62}\n\n", "dim")

    def _clear(self):
        self.chat.config(state="normal")
        self.chat.delete("1.0", "end")
        self.chat.config(state="disabled")
        self.session_results = None
        self._reset_sidebar()

    def _reset_sidebar(self):
        self._lbl_ticker.config(text="—", fg=ACCENT)
        self._lbl_num.config(text="—", fg=FG_DIM)
        self._lbl_sig.config(text="—", fg=FG_DIM)
        self._lbl_fit.config(text="")
        for w in self._comp_frame.winfo_children():
            w.destroy()

    def _update_sidebar(self, composite):
        sig_color = {
            "STRONG BUY": GREEN, "BUY": GREEN,
            "WATCHLIST": YELLOW, "AVOID": RED,
        }.get(composite.signal, FG)

        self._lbl_ticker.config(text=composite.ticker)
        self._lbl_num.config(text=f"{composite.total_score:.0f}", fg=sig_color)
        self._lbl_sig.config(text=composite.signal, fg=sig_color)
        self._lbl_fit.config(text=composite.account_fit)

        for w in self._comp_frame.winfo_children():
            w.destroy()

        for c in composite.components:
            row = tk.Frame(self._comp_frame, bg=BG2)
            row.pack(fill="x", pady=1)
            short = c.name.split("—")[1].strip() if "—" in c.name else c.name[:16]
            tk.Label(row, text=short[:18], font=FONT_SM, bg=BG2, fg=FG_DIM,
                     width=16, anchor="w").pack(side="left")
            canvas = tk.Canvas(row, width=56, height=8, bg=BG3,
                                highlightthickness=0)
            canvas.pack(side="left", padx=2)
            fill_w = max(2, int(c.raw * 54))
            bar_color = GREEN if c.raw >= 0.65 else (YELLOW if c.raw >= 0.40 else RED)
            canvas.create_rectangle(0, 0, fill_w, 8, fill=bar_color, outline="")
            tk.Label(row, text=f"{c.raw:.0%}", font=FONT_SM, bg=BG2, fg=FG_DIM,
                     width=4).pack(side="left")

    # ── PORTFOLIO ───────────────────────────
    def _reload_portfolio(self):
        path = filedialog.askopenfilename(
            title="Select portfolio.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        self.port_label.config(text="📡 Fetching prices...", fg=FG_DIM)
        def _reload():
            fetch_live_prices(path)
            self.portfolio_ctx = load_portfolio_context(path)
            self.port_label.config(text="📋 Portfolio loaded", fg=GREEN)
        threading.Thread(target=_reload, daemon=True).start()

    # ── ANALYSIS ────────────────────────────
    def _toggle_run(self):
        if self.is_running:
            self.stop_requested = True
            self.run_btn.config(text="⏹ Stopping...", bg=RED, fg="#000000", state="disabled")
        else:
            ticker = self.ticker_var.get().strip()
            if not ticker:
                self.ticker_entry.focus()
                return
            self._start_analysis(ticker)

    def _start_analysis(self, ticker):
        self.is_running      = True
        self.stop_requested  = False
        self.session_results = None
        self.run_btn.config(bg=RED, fg="#000000", text="⏹ Stop")
        self.ticker_entry.config(state="disabled")
        self.qa_entry.config(state="disabled", fg=FG_DIM)
        self.qa_btn.config(state="disabled", bg=BG3, fg=FG_DIM)
        mode = self.mode_var.get()
        threading.Thread(target=self._analysis_thread, args=(ticker, mode), daemon=True).start()

    def _analysis_thread(self, ticker, mode="composite"):
        """
        mode="composite" → fast deterministic score only
        mode="long"      → score + Claude narrative for each framework
        """
        try:
            self._w(f"\n")
            self._w(f"ANALYZING  {ticker.upper()}\n", "header")
            self._rule()

            stages = [
                "Resolving ticker...", "Fetching live data...",
                "Buffett analysis...", "Weiss analysis...",
                "Bogle analysis...", "Dalio analysis...",
                "Druckenmiller analysis...", "Building composite score...",
            ]
            for s in stages:
                if self.stop_requested:
                    self._w("  ⏹ Stopped.\n", "dim")
                    return
                self.status_lbl.config(text=s)

            results, err = run_full_analysis(ticker, PORTFOLIO_FILE)
            if err:
                self._w(f"❌ {err}\n", "red")
                return
            if self.stop_requested:
                self._w("  ⏹ Stopped.\n", "dim")
                return

            self.session_results = results
            composite = results["composite"]
            live = results.get("live_data")

            # Company header
            asset_class = results.get("asset_class", "UNKNOWN")
            asset_note  = results.get("asset_note", "")
            asset_tag = {
                "ETF": "yellow", "REIT": "teal",
                "STOCK": "green", "MUTUAL_FUND": "yellow"
            }.get(asset_class, "dim")

            if live:
                price  = f"${live.current_price:.2f}" if live.current_price else "N/A"
                sector = live.sector or "N/A"
                mktcap = f"${live.market_cap/1e9:.1f}B" if live.market_cap else "N/A"
                self._w(f"  {composite.ticker} — {composite.company_name}\n", "blue")
                self._w(f"  Price: {price}  |  Sector: {sector}  |  MCap: {mktcap}  |  ", "dim")
                self._w(f"Type: {asset_class}\n", asset_tag)
            if asset_note:
                self._w(f"  ℹ️  {asset_note}\n", "yellow")
            # Show which analyzers were skipped
            skipped = [k.replace("_skipped","") for k in results
                       if k.endswith("_skipped") and results[k]]
            if skipped:
                self._w(f"  ⏭  Skipped (not applicable): {', '.join(skipped)}\n", "dim")
            self._w(f"  {composite.market_context}\n\n", "dim")

            # Score table
            self._write_score_table(composite)

            # Framework detail
            self._write_framework_details(composite, live, results)

            # Sidebar
            self._update_sidebar(composite)

            # Long mode: Claude narrative per framework
            if mode == "long" and not self.stop_requested:
                self._write_long_narratives(results)

            # Enable Q&A
            self.qa_entry.config(state="normal", fg=FG)
            self.qa_btn.config(state="normal", bg="#238636", fg="#FFFFFF")
            self._w(f"\n  💬 Ask Claude about this analysis below\n", "dim")
            self._rule()

            self.status_lbl.config(text=f"Done — {composite.total_score:.0f}/100  {composite.signal}")

        except Exception as e:
            self._w(f"\n❌ Error: {e}\n", "red")
            self._w(traceback.format_exc(), "dim")
            self.status_lbl.config(text="Error")
        finally:
            self.is_running = False
            self.stop_requested = False
            self.run_btn.config(state="normal", text="▶  Analyze", bg=ACCENT, fg="#000000")
            self.ticker_entry.config(state="normal")
            self.ticker_var.set("")
            self.ticker_entry.focus()

    def _write_score_table(self, composite):
        sig_tag = {
            "STRONG BUY": "score_strong", "BUY": "score_buy",
            "WATCHLIST": "score_watch", "AVOID": "score_avoid",
        }.get(composite.signal, "dim")

        self._w("  COMPOSITE SCORE\n", "header")
        self._w(f"  {'─'*58}\n", "dim")
        self._w(f"  {'Framework':<30} {'Score':>6}  {'Wt':>3}  Bar\n", "dim")
        self._w(f"  {'─'*58}\n", "dim")

        for c in composite.components:
            bar = "█" * c.bar_filled + "░" * (20 - c.bar_filled)
            bar_tag = "green" if c.raw >= 0.65 else ("yellow" if c.raw >= 0.40 else "red")
            self._w(f"  {c.name:<30} {c.raw:>6.0%}  {c.weight:>3}%  ", "dim")
            self._w(f"{bar}\n", bar_tag)

        self._w(f"  {'─'*58}\n", "dim")
        total_bar = "█" * round(composite.total_score/5) + "░" * (20 - round(composite.total_score/5))
        self._w(f"  {'TOTAL SCORE':<30} {composite.total_score:>6.1f}  100%  ", "dim")
        self._w(f"{total_bar}\n\n", sig_tag)

        self._w(f"  SIGNAL:  ", "dim")
        self._w(f"{composite.signal}\n", sig_tag)
        self._w(f"  ACCOUNT: {composite.account_fit}\n", "blue")
        self._w(f"  REASON:  {composite.account_reason}\n", "dim")
        self._w(f"  DATA:    {composite.data_quality} quality", "dim")
        if composite.missing_data:
            self._w(f"  |  Missing: {', '.join(composite.missing_data)}", "dim")
        self._w("\n\n")

    def _write_framework_details(self, composite, live, results):
        """Show full formatted breakdown from each analyzer, not just the summary score."""

        # ── BUFFETT ──────────────────────────────────────────────────────
        buffett = results.get("buffett")
        if buffett and not results.get("buffett_skipped"):
            self._rule("🎩 Buffett — Moat & Valuation")
            try:
                from buffett_analyzer import format_display_summary
                self._w(format_display_summary(buffett) + "\n", "dim")
            except Exception as e:
                self._w(f"  Display error: {e}\n", "red")

        # ── WEISS ─────────────────────────────────────────────────────────
        weiss = results.get("weiss")
        if weiss and not results.get("weiss_skipped"):
            self._rule("📈 Weiss — Yield Signal & Blue Chip Criteria")
            try:
                from weiss_analyzer import format_weiss_display
                self._w(format_weiss_display(weiss) + "\n", "dim")
            except Exception as e:
                self._w(f"  Display error: {e}\n", "red")

        # ── BOGLE ─────────────────────────────────────────────────────────
        bogle = results.get("bogle")
        if bogle and not results.get("bogle_skipped"):
            self._rule("📊 Bogle — Past Performance, Timing & Diversification")
            try:
                from bogle_analyzer import format_bogle_display
                self._w(format_bogle_display(bogle) + "\n", "dim")
            except Exception as e:
                self._w(f"  Display error: {e}\n", "red")

        # ── DALIO ─────────────────────────────────────────────────────────
        dalio = results.get("dalio")
        if dalio and not results.get("dalio_skipped"):
            self._rule("🌊 Dalio — Four Filters")
            try:
                from dalio_analyzer import format_dalio_display
                self._w(format_dalio_display(dalio) + "\n", "dim")
            except Exception as e:
                self._w(f"  Display error: {e}\n", "red")

        # ── DRUCKENMILLER ─────────────────────────────────────────────────
        druck = results.get("druckenmiller")
        if druck and not results.get("druckenmiller_skipped"):
            self._rule("📡 Druckenmiller — Five Pillars")
            try:
                from druckenmiller_analyzer import format_druckenmiller_display
                self._w(format_druckenmiller_display(druck) + "\n", "dim")
            except Exception as e:
                self._w(f"  Display error: {e}\n", "red")

        # ── CHAMATH ───────────────────────────────────────────────────────────
        chamath = results.get("chamath")
        if chamath and not results.get("chamath_skipped"):
            self._rule("💰 Chamath Palihapitiya — Narrative Squeeze Analysis")
            try:
                from squeeze_analyzers import format_chamath_display
                self._w(format_chamath_display(chamath) + "\n", "dim")
            except Exception as e:
                self._w(f"  Display error: {e}\n", "red")

        # ── RAW KEY METRICS ────────────────────────────────────────────────
        if live:
            self._rule("📋 Key Metrics")
            buffett_obj = results.get("buffett")
            roic = f"{buffett_obj.moat.roic:.1%}" if buffett_obj and buffett_obj.moat.roic else "N/A"

            def fmt(val, style="num"):
                if val is None: return "N/A"
                if style == "pct":  return f"{val:.1%}"
                if style == "pct2": return f"{val:.2%}"
                if style == "x":    return f"{val:.1f}x"
                if style == "bil":  return f"${val/1e9:.2f}B"
                if style == "dol":  return f"${val:.2f}"
                return f"{val:.2f}"

            rows = [
                ("ROIC",             roic,                                     ""),
                ("Gross Margin",     fmt(live.gross_margin, "pct"),            "> 40% (Buffett)"),
                ("Debt / Equity",    fmt(live.debt_to_equity),                 "< 0.5 (Buffett)"),
                ("Free Cash Flow",   fmt(live.free_cash_flow, "bil"),          ""),
                ("P/E (Trailing)",   fmt(live.pe_ratio, "x"),                  "< 20 (Weiss)"),
                ("PEG Ratio",        fmt(live.peg_ratio),                      "< 1.0 best (Lynch)"),
                ("Dividend Yield",   fmt(live.dividend_yield, "pct2") if live.dividend_yield else "None", ""),
                ("Payout Ratio",     fmt(live.payout_ratio, "pct") if live.payout_ratio else "N/A", "< 50% (Weiss)"),
                ("Beta",             fmt(live.beta),                           "1.0 = market"),
                ("52wk High",        fmt(live.fifty_two_wk_high, "dol"),       ""),
                ("52wk Low",         fmt(live.fifty_two_wk_low, "dol"),        ""),
                ("Earnings Growth",  fmt(live.earnings_growth, "pct") if live.earnings_growth else "N/A", ""),
                ("Revenue Growth",   fmt(live.revenue_growth, "pct") if live.revenue_growth else "N/A",  ""),
            ]
            self._w(f"  {'Metric':<24} {'Value':>10}   Context\n", "header")
            self._w(f"  {'─'*58}\n", "dim")
            for label, value, ctx in rows:
                self._w(f"  {label:<24} {value:>10}   {ctx}\n", "dim")
            self._w("\n")

    # ── CLAUDE Q&A ───────────────────────────
    def _write_long_narratives(self, results):
        """Long mode: ask Claude for a narrative on each framework section."""
        from composite_score import format_composite_for_claude
        composite = results["composite"]
        ctx = format_composite_for_claude(composite)

        self._rule("Deep Dive — Claude Narratives")
        self._w("  Each framework interpreted by Claude based on calculated scores.\n\n", "dim")

        framework_prompts = [
            ("Buffett — Moat & Valuation",
             "Based on the Buffett Moat and Valuation scores shown, give a 3-sentence assessment. "
             "What does the ROIC tell you about the moat? Is the DCF/valuation compelling? "
             "Reference the actual score numbers."),
            ("Weiss — Yield & Blue Chip Quality",
             "Based on the Weiss Yield Signal and Blue Chip Quality scores, give a 3-sentence assessment. "
             "Is the yield signal actionable? How many of the 7 criteria does this pass? "
             "Reference the actual scores."),
            ("Bogle — Timing & Diversification",
             "Based on the Bogle Buy Timing and Diversification scores, give a 3-sentence assessment. "
             "Is now a good entry point by mean reversion? Does adding this help or hurt the portfolio? "
             "Reference the actual scores."),
            ("Dalio — Debt & Bubble Risk",
             "Based on the Dalio Debt Cycle and Bubble Risk scores, give a 3-sentence assessment. "
             "Is the balance sheet safe in a high-rate environment? Any bubble signals? "
             "Reference the actual scores."),
            ("Lynch — PEG & Growth Classification",
             "Based on the Lynch PEG score, give a 3-sentence assessment. "
             "Classify this company (slow grower, stalwart, fast grower, cyclical, turnaround). "
             "Is the growth priced in fairly? Reference the actual score."),
            ("Druckenmiller — Triple Alignment & Macro",
             f"Based on the Druckenmiller Triple Alignment score shown, give a 3-sentence assessment. "
             f"Is the liquidity regime favorable? Is momentum accelerating? Does the chart confirm? "
             f"What conviction level would Druckenmiller assign and why?"),
            ("Portfolio Fit & Recommendation",
             f"Given the overall composite score of {composite.total_score:.1f}/100 and signal of {composite.signal}, "
             f"give a final 3-sentence recommendation for Johnathan. "
             f"Which account (Roth 401k or taxable brokerage) fits best and why? "
             f"Should he deploy any of his $9,000 cash here?"),
        ]

        for section_name, prompt in framework_prompts:
            if self.stop_requested:
                self._w("  ⏹ Stopped.\n", "dim")
                return

            self._w(f"  {section_name}\n", "header")
            self._w(f"  {'─'*54}\n", "dim")
            self._w("  ⏳ Asking Claude...\n", "dim")

            full_prompt = f"{ctx}\n\nQuestion: {prompt}"
            answer = ask_claude(full_prompt, ctx, self.portfolio_ctx)

            self.chat.config(state="normal")
            pos = self.chat.search("  ⏳ Asking Claude...\n", "1.0", "end")
            if pos:
                self.chat.delete(pos, f"{pos}+{len('  ⏳ Asking Claude...')+1}c")
            self.chat.config(state="disabled")

            self._w(f"  {answer}\n\n", "claude")

    def _ask_claude(self):
        if self.is_running or not self.session_results:
            return
        question = self.qa_var.get().strip()
        if not question:
            self.qa_entry.focus()
            return
        self.is_running = True
        self.qa_btn.config(state="disabled", text="⏳")
        self.qa_entry.config(state="disabled")
        self.run_btn.config(state="disabled")
        threading.Thread(target=self._claude_thread, args=(question,), daemon=True).start()

    def _claude_thread(self, question):
        try:
            from composite_score import format_composite_for_claude
            ctx = format_composite_for_claude(self.session_results["composite"])
            self._rule("Claude Q&A")
            self._w(f"  Q: {question}\n\n", "blue")
            self._w("  ⏳ Thinking...\n", "dim")
            answer = ask_claude(question, ctx, self.portfolio_ctx)
            self.chat.config(state="normal")
            pos = self.chat.search("  ⏳ Thinking...\n", "1.0", "end")
            if pos:
                self.chat.delete(pos, f"{pos}+{len('  ⏳ Thinking...')+1}c")
            self.chat.config(state="disabled")
            self._w(f"  {answer}\n\n", "claude")
        finally:
            self.is_running = False
            self.qa_btn.config(state="normal", text="💬 Ask", bg="#238636", fg="#FFFFFF")
            self.qa_entry.config(state="normal", fg=FG)
            self.run_btn.config(state="normal")
            self.qa_var.set("")
            self.qa_entry.focus()


if __name__ == "__main__":
    root = tk.Tk()
    app = RoundtableApp(root)
    root.mainloop()
