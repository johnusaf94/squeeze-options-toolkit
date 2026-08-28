"""
portfolio_builder_gui.py
=========================
Standalone portfolio builder.
Requires: shared_utils.py, portfolio_builder.py, ticker_resolver.py
"""

# ── GLOBAL yfinance RATE LIMITER ────────────────────────────────
# Must be imported BEFORE anything that uses yfinance. Monkey-
# patches yfinance.Ticker with token-bucket rate limiting + caching.
import yfinance_throttle  # noqa: F401  # installs global throttle


import tkinter as tk
from tkinter import scrolledtext, filedialog
import threading
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from shared_utils import *

PORTFOLIO_FILE = "portfolio.xlsx"

def fetch_live_prices(filepath):
    try:
        import openpyxl, yfinance as yf
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Holdings"]
        tickers = []
        row_map = {}
        for row in ws.iter_rows(min_row=5, values_only=False):
            t = row[0].value
            if not t or str(t).strip() in ("", "TOTALS", "CASH"):
                continue
            t = str(t).strip().upper()
            tickers.append(t)
            row_map[t] = row[4].row
        if tickers:
            data = yf.download(tickers, period="1d", progress=False, auto_adjust=True)
            close = data["Close"]
            thin = Side(style="thin", color="30363D")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for t in tickers:
                try:
                    price = float(close[t].iloc[-1]) if len(tickers) > 1 else float(close.iloc[-1])
                    c = ws.cell(row=row_map[t], column=5, value=round(price, 4))
                    c.font = Font(name="Consolas", size=10, color="007700", bold=True)
                    c.fill = PatternFill("solid", start_color="E8FFE8")
                    c.number_format = "$#,##0.00"
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.border = border
                except Exception:
                    pass
        wb.save(filepath)
    except Exception:
        pass


# ─────────────────────────────────────────────
# ANALYSIS PIPELINE
# ─────────────────────────────────────────────


class PortfolioBuilderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🏗 Portfolio Builder")
        self.root.geometry("1280x860")
        self.root.configure(bg=BG)
        self._pb_running = False
        self._pb_stop    = False
        self.portfolio_ctx = load_portfolio_context(PORTFOLIO_FILE)

        top = tk.Frame(root, bg=BG2, pady=8)
        top.pack(fill="x")
        tk.Label(top, text="🏗 PORTFOLIO BUILDER", font=FONT_HD,
                 bg=BG2, fg=ACCENT).pack(side="left", padx=16)
        self.port_label = tk.Label(top, text="📋 Loading...", font=FONT_SM, bg=BG2, fg=FG_DIM)
        self.port_label.pack(side="right", padx=8)
        tk.Button(top, text="Reload Portfolio", font=FONT_SM, bg=BG3, fg=FG,
                  relief="flat", cursor="hand2",
                  command=self._reload_portfolio).pack(side="right", padx=4)

        self.tab_portfolio = tk.Frame(root, bg=BG)
        self.tab_portfolio.pack(fill="both", expand=True)
        self._build_portfolio_tab()
        threading.Thread(target=self._init_prices, daemon=True).start()

    def _build_portfolio_tab(self):
        """Portfolio Builder — Michael's recursive discovery engine."""
        import tkinter.ttk as ttk
        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            HAS_MATPLOTLIB = True
        except ImportError:
            HAS_MATPLOTLIB = False

        parent = self.tab_portfolio
        self._pb_running   = False
        self._pb_stop      = False
        self._pb_portfolio = None
        self._pb_candidates = []

        # ── LEFT: Scrollable Controls ────────────────────────────────────
        # Wrap in a canvas so it scrolls when content exceeds window height
        ctrl_outer = tk.Frame(parent, bg=BG2, width=290)
        ctrl_outer.pack(side="left", fill="y")
        ctrl_outer.pack_propagate(False)

        ctrl_canvas = tk.Canvas(ctrl_outer, bg=BG2, width=270,
                                highlightthickness=0, bd=0)
        ctrl_scrollbar = tk.Scrollbar(ctrl_outer, orient="vertical",
                                       command=ctrl_canvas.yview)
        ctrl_canvas.configure(yscrollcommand=ctrl_scrollbar.set)

        ctrl_scrollbar.pack(side="right", fill="y")
        ctrl_canvas.pack(side="left", fill="both", expand=True)

        ctrl = tk.Frame(ctrl_canvas, bg=BG2)
        ctrl_window = ctrl_canvas.create_window((0, 0), window=ctrl, anchor="nw")

        def _on_ctrl_configure(e):
            ctrl_canvas.configure(scrollregion=ctrl_canvas.bbox("all"))
            ctrl_canvas.itemconfig(ctrl_window, width=ctrl_canvas.winfo_width())

        ctrl.bind("<Configure>", _on_ctrl_configure)
        ctrl_canvas.bind("<Configure>", lambda e: ctrl_canvas.itemconfig(
            ctrl_window, width=e.width))

        # Mouse wheel scrolling
        def _on_mousewheel(e):
            ctrl_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        ctrl_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        tk.Label(ctrl, text="PORTFOLIO BUILDER", font=FONT_LG,
                 bg=BG2, fg=ACCENT).pack(pady=(14,4), padx=14, anchor="w")
        tk.Label(ctrl, text="Portfolio Builder — Composite Score Engine",
                 font=FONT_SM, bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=10)

        # ── Inputs ──
        def labeled_entry(parent, label, default, is_int=False):
            f = tk.Frame(parent, bg=BG2)
            f.pack(fill="x", padx=14, pady=3)
            tk.Label(f, text=label, font=FONT_SM, bg=BG2, fg=FG_DIM,
                     width=18, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(default))
            tk.Entry(f, textvariable=var, font=FONT_SM, bg=BG3, fg=FG,
                     insertbackground=FG, relief="flat", bd=4, width=10).pack(side="left")
            return var

        self._pb_cash   = labeled_entry(ctrl, "Starting Cash ($)",     "10000")
        self._pb_years  = labeled_entry(ctrl, "Years to Retirement",   "30")
        self._pb_monthly= labeled_entry(ctrl, "Monthly Contrib ($)",   "600")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Sector selector ──
        tk.Label(ctrl, text="SECTORS TO SCAN", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        from portfolio_builder import SECTOR_UNIVERSE
        self._pb_sectors = {}
        sectors_frame = tk.Frame(ctrl, bg=BG2)
        sectors_frame.pack(fill="x", padx=10)

        all_sectors = list(SECTOR_UNIVERSE.keys())
        # Default select most growth-oriented sectors
        default_on = {"Technology","Healthcare","Consumer Discretionary",
                      "Financials","Industrials","Communication Services"}

        for i, sec in enumerate(all_sectors):
            var = tk.BooleanVar(value=sec in default_on)
            self._pb_sectors[sec] = var
            short = sec[:22]
            tk.Checkbutton(sectors_frame, text=short, variable=var,
                           font=FONT_SM, bg=BG2, fg=FG, selectcolor=BG3,
                           activebackground=BG2, relief="flat").pack(anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Agent selector ──
        tk.Label(ctrl, text="ANALYST AGENTS", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        self._pb_agents = {}
        agents = [
            ("buffett",        "🎩 Buffett"),
            ("weiss",          "📈 Weiss"),
            ("bogle",          "📊 Bogle"),
            ("dalio",          "🌊 Dalio"),
            ("druckenmiller",  "📡 Druckenmiller"),
        ]
        agents_frame = tk.Frame(ctrl, bg=BG2)
        agents_frame.pack(fill="x", padx=10)
        for key, label in agents:
            var = tk.BooleanVar(value=True)
            self._pb_agents[key] = var
            tk.Checkbutton(agents_frame, text=label, variable=var,
                           font=FONT_SM, bg=BG2, fg=FG, selectcolor=BG3,
                           activebackground=BG2, relief="flat").pack(anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Portfolio Constraints ──
        tk.Label(ctrl, text="PORTFOLIO CONSTRAINTS", font=FONT_SM,
                 bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        # Tickers per sector
        f_tps = tk.Frame(ctrl, bg=BG2)
        f_tps.pack(fill="x", padx=14, pady=2)
        tk.Label(f_tps, text="Tickers/sector", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_tickers_per = tk.StringVar(value="5")
        tk.Entry(f_tps, textvariable=self._pb_tickers_per, font=FONT_SM,
                 bg=BG3, fg=FG, relief="flat", bd=4, width=6).pack(side="left")

        # Max sector allocation slider
        f_sec = tk.Frame(ctrl, bg=BG2)
        f_sec.pack(fill="x", padx=14, pady=4)
        tk.Label(f_sec, text="Max Sector %", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_max_sector = tk.IntVar(value=35)
        self._pb_sector_lbl = tk.Label(f_sec, text="35%", font=FONT_SM,
                                        bg=BG2, fg=ACCENT, width=5)
        self._pb_sector_lbl.pack(side="right")

        def _update_sector_lbl(val):
            self._pb_sector_lbl.config(text=f"{int(float(val))}%")

        tk.Scale(ctrl, from_=10, to=100, orient="horizontal",
                 variable=self._pb_max_sector,
                 command=_update_sector_lbl,
                 bg=BG2, fg=FG, troughcolor=BG3,
                 highlightthickness=0, bd=0,
                 sliderlength=16, showvalue=False,
                 length=220).pack(padx=14, pady=(0,4))

        # Target return slider
        f_ret = tk.Frame(ctrl, bg=BG2)
        f_ret.pack(fill="x", padx=14, pady=2)
        tk.Label(f_ret, text="Target Return %", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_target_return = tk.IntVar(value=15)
        self._pb_return_lbl = tk.Label(f_ret, text="15%", font=FONT_SM,
                                        bg=BG2, fg=GREEN, width=5)
        self._pb_return_lbl.pack(side="right")

        def _update_return_lbl(val):
            v = int(float(val))
            color = GREEN if v <= 15 else (YELLOW if v <= 20 else RED)
            self._pb_return_lbl.config(text=f"{v}%", fg=color)

        tk.Scale(ctrl, from_=5, to=30, orient="horizontal",
                 variable=self._pb_target_return,
                 command=_update_return_lbl,
                 bg=BG2, fg=FG, troughcolor=BG3,
                 highlightthickness=0, bd=0,
                 sliderlength=16, showvalue=False,
                 length=220).pack(padx=14, pady=(0,6))

        # Max position per stock slider
        f_pos = tk.Frame(ctrl, bg=BG2)
        f_pos.pack(fill="x", padx=14, pady=4)
        tk.Label(f_pos, text="Max Position %", font=FONT_SM,
                 bg=BG2, fg=FG_DIM, width=18, anchor="w").pack(side="left")
        self._pb_max_position = tk.IntVar(value=20)
        self._pb_position_lbl = tk.Label(f_pos, text="20%", font=FONT_SM,
                                          bg=BG2, fg=ACCENT, width=5)
        self._pb_position_lbl.pack(side="right")

        def _update_position_lbl(val):
            self._pb_position_lbl.config(text=f"{int(float(val))}%")

        tk.Scale(ctrl, from_=5, to=50, orient="horizontal",
                 variable=self._pb_max_position,
                 command=_update_position_lbl,
                 bg=BG2, fg=FG, troughcolor=BG3,
                 highlightthickness=0, bd=0,
                 sliderlength=16, showvalue=False,
                 length=220).pack(padx=14, pady=(0,6))

        tk.Label(ctrl, text="5% = highly diversified  |  50% = concentrated",
                 font=("Consolas",7), bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w")

        tk.Label(ctrl, text="5% = conservative  |  30% = aggressive",
                 font=("Consolas",7), bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w")

        tk.Frame(ctrl, bg=BORDER, height=1).pack(fill="x", padx=10, pady=8)

        # ── Include ETFs toggle ──
        self._pb_include_etf = tk.BooleanVar(value=False)
        etf_frame = tk.Frame(ctrl, bg=BG2)
        etf_frame.pack(fill="x", padx=10, pady=2)

        self._pb_etf_btn = tk.Button(
            etf_frame,
            text="☐  Include Sector ETFs",
            font=FONT_SM,
            bg=BG3, fg=FG_DIM,
            relief="flat", cursor="hand2",
            anchor="w", padx=10, pady=5,
            command=self._pb_toggle_etf,
        )
        self._pb_etf_btn.pack(fill="x")
        tk.Label(ctrl, text="Adds top ETFs by market cap per sector",
                 font=("Consolas",7), bg=BG2, fg=FG_DIM).pack(padx=14, anchor="w", pady=(0,4))

        # Run button
        self._pb_run_btn = tk.Button(ctrl, text="▶  Build Portfolio",
                                      font=("Consolas",11,"bold"),
                                      bg=ACCENT, fg="#000000",
                                      relief="flat", cursor="hand2",
                                      padx=14, pady=6,
                                      command=self._pb_toggle)
        self._pb_run_btn.pack(fill="x", padx=10, pady=8)

        self._pb_status = tk.Label(ctrl, text="Ready", font=FONT_SM,
                                    bg=BG2, fg=FG_DIM, wraplength=240)
        self._pb_status.pack(padx=14, anchor="w")

        # ── RIGHT: Output split ──────────────────────────────────────────
        right = tk.Frame(parent, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Top: pie chart area
        chart_frame = tk.Frame(right, bg=BG, height=320)
        chart_frame.pack(fill="x", pady=0)
        chart_frame.pack_propagate(False)

        if HAS_MATPLOTLIB:
            fig = Figure(figsize=(5, 3.2), dpi=90, facecolor=BG)
            fig.subplots_adjust(left=0.02, right=0.98, top=0.92, bottom=0.02)
            self._pb_fig = fig
            self._pb_ax  = fig.add_subplot(111)
            self._pb_ax.set_facecolor(BG)
            self._pb_canvas = FigureCanvasTkAgg(fig, master=chart_frame)
            self._pb_canvas.get_tk_widget().pack(fill="both", expand=True)
            self._draw_empty_pie()
        else:
            self._pb_canvas = None
            self._pb_fig    = None
            tk.Label(chart_frame,
                     text="Install matplotlib for pie chart: pip install matplotlib",
                     font=FONT_SM, bg=BG, fg=YELLOW).pack(pady=40)

        # Bottom: log + results
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x")
        self._pb_log = scrolledtext.ScrolledText(
            right, wrap="word", font=FONT_SM, bg=BG, fg=FG,
            insertbackground=FG, relief="flat", borderwidth=0,
            state="disabled", padx=16, pady=10,
        )
        self._pb_log.pack(fill="both", expand=True)

        # Log tags
        for tag, cfg in [
            ("header",  {"font": FONT_LG,               "foreground": ACCENT}),
            ("dim",     {                                 "foreground": FG_DIM}),
            ("green",   {                                 "foreground": GREEN}),
            ("red",     {                                 "foreground": RED}),
            ("yellow",  {                                 "foreground": YELLOW}),
            ("blue",    {                                 "foreground": BLUE}),
            ("accent_teal", {"font": ("Consolas",10,"bold"), "foreground": TEAL}),
            ("include", {"font": ("Consolas",9,"bold"),  "foreground": GREEN}),
            ("reject",  {"font": ("Consolas",9),         "foreground": RED}),
            ("watch",   {"font": ("Consolas",9),         "foreground": YELLOW}),
        ]:
            self._pb_log.tag_config(tag, **cfg)


    def _pb_log_write(self, text, tag=None):
        self._pb_log.config(state="normal")
        self._pb_log.insert("end", text, tag) if tag else self._pb_log.insert("end", text)
        self._pb_log.see("end")
        self._pb_log.config(state="disabled")
        self.root.update_idletasks()


    def _pb_toggle_etf(self):
        """Toggle Include ETFs button state."""
        current = self._pb_include_etf.get()
        self._pb_include_etf.set(not current)
        if self._pb_include_etf.get():
            self._pb_etf_btn.config(
                text="☑  Include Sector ETFs",
                bg="#238636", fg="#FFFFFF"
            )
        else:
            self._pb_etf_btn.config(
                text="☐  Include Sector ETFs",
                bg=BG3, fg=FG_DIM
            )


    def _pb_toggle(self):
        if self._pb_running:
            self._pb_stop = True
            self._pb_run_btn.config(text="⏹ Stopping...", bg=RED, state="disabled")
        else:
            self._pb_start()


    def _pb_start(self):
        try:
            cash       = float(self._pb_cash.get().replace(",",""))
            years      = int(self._pb_years.get())
            monthly    = float(self._pb_monthly.get().replace(",",""))
            tps        = int(self._pb_tickers_per.get())
            max_sector   = self._pb_max_sector.get()   / 100.0
            max_position = self._pb_max_position.get() / 100.0
            target_ret   = self._pb_target_return.get() / 100.0
            include_etf  = self._pb_include_etf.get()
        except ValueError:
            self._pb_log_write("❌ Invalid inputs — check Starting Cash, Years, Monthly Contrib.\n", "red")
            return

        sectors = [s for s, v in self._pb_sectors.items() if v.get()]
        agents  = [a for a, v in self._pb_agents.items() if v.get()]

        if not sectors:
            self._pb_log_write("❌ Select at least one sector.\n", "red")
            return
        if not agents:
            self._pb_log_write("❌ Select at least one analyst agent.\n", "red")
            return

        self._pb_running = True
        self._pb_stop    = False
        self._pb_run_btn.config(bg=RED, fg="#000000", text="⏹ Stop")
        self._draw_empty_pie()

        threading.Thread(
            target=self._pb_thread,
            args=(cash, years, monthly, sectors, agents, tps,
                  max_sector, max_position, target_ret, include_etf),
            daemon=True,
        ).start()


    def _pb_thread(self, cash, years, monthly, sectors, agents, tps,
                    max_sector=0.35, max_position=0.20, target_ret=0.15, include_etf=False):
        try:
            self._pb_log.config(state="normal")
            self._pb_log.delete("1.0", "end")
            self._pb_log.config(state="disabled")

            self._pb_log_write("\n")
            self._pb_log_write("  PORTFOLIO BUILDER — DISCOVERY RUN\n", "header")
            self._pb_log_write(f"  Cash: ${cash:,.0f}  |  Horizon: {years}yr  |  Contrib: ${monthly:,.0f}/mo\n", "dim")
            self._pb_log_write(f"  Target Return: {target_ret:.0%}  |  Max Sector: {max_sector:.0%}  |  Max Position: {max_position:.0%}  |  ETFs: {'YES' if include_etf else 'NO'}\n", "dim")
            self._pb_log_write(f"  Scanning {len(sectors)} sectors × {tps} tickers = "
                                f"up to {len(sectors)*tps} candidates\n", "dim")
            self._pb_log_write(f"  Agents: {', '.join(agents)}\n\n", "dim")

            self._pb_status.config(text="Scanning sectors...")

            from portfolio_builder import run_portfolio_builder

            def progress(ticker, msg):
                if self._pb_stop:
                    return
                # Color by score label (no binary gate anymore)
                tag = "dim"
                if "STRONG" in msg:
                    tag = "include"
                elif "GOOD" in msg:
                    tag = "green"
                elif "FAIR" in msg:
                    tag = "yellow"
                elif "WEAK" in msg:
                    tag = "reject"
                elif "⏭" in msg:
                    tag = "dim"
                self._pb_log_write(f"  {msg}\n", tag)
                self._pb_status.config(text=f"Scanning {ticker}...")

            portfolio, candidates = run_portfolio_builder(
                starting_cash=cash,
                years=years,
                monthly_contrib=monthly,
                selected_sectors=sectors,
                selected_agents=agents,
                tickers_per_sector=tps,
                max_sector_pct=max_sector,
                max_position_pct=max_position,
                target_annual_return=target_ret,
                include_etf=include_etf,
                progress_callback=progress,
            )

            if self._pb_stop:
                self._pb_log_write("\n  ⏹ Stopped.\n", "dim")
                return

            self._pb_portfolio  = portfolio
            self._pb_candidates = candidates

            # ── Michael's results ──
            self._pb_log_write("\n")
            self._pb_log_write("  ─" * 30 + "\n", "dim")
            self._pb_log_write("  PORTFOLIO BUILDER RESULTS\n", "header")
            self._pb_log_write("  ─" * 30 + "\n\n", "dim")

            self._pb_log_write(f"  {portfolio.pb_summary}\n\n", "include")

            # Position table
            self._pb_log_write(
                f"  {'Ticker':<8} {'Company':<24} {'Sector':<22} {'Alloc':>6}  "
                f"{'Return':>7}  {'Sharpe':>6}  {'Vol':>7}  {'VolScore':>8}  {'Comp':>5}\n", "blue"
            )
            self._pb_log_write(f"  {'─'*92}\n", "dim")

            # Pull volatility from candidates for display
            cand_map = {c.ticker: c for c in candidates}
            for pos in portfolio.positions:
                ret_s  = f"{pos.annual_return:.1%}"  if pos.annual_return  else "N/A"
                sha_s  = f"{pos.sharpe_ratio:.2f}"   if pos.sharpe_ratio   else "N/A"
                cand   = cand_map.get(pos.ticker)
                vol_s  = f"{cand.annual_volatility:.1%}" if cand and cand.annual_volatility else "N/A"
                pb_s   = f"{cand.pb_score:.0f}"      if cand else "N/A"
                # Color by pb_score
                tag = "green" if (cand and cand.pb_score >= 65) else ("yellow" if (cand and cand.pb_score >= 45) else "dim")
                self._pb_log_write(
                    f"  {pos.ticker:<8} {pos.company_name[:24]:<24} "
                    f"{pos.sector[:22]:<22} {pos.allocation_pct:>5.1f}%  "
                    f"{ret_s:>7}  {sha_s:>6}  {vol_s:>7}  {pb_s:>8}  {pos.composite_score:>5.1f}\n",
                    tag
                )

            self._pb_log_write(f"\n  Cash remaining: ${portfolio.cash_remaining:,.0f}\n", "dim")

            # Sector breakdown
            self._pb_log_write("\n  SECTOR ALLOCATION\n", "blue")
            for sec, pct in sorted(portfolio.sector_weights.items(), key=lambda x: -x[1]):
                bar = "█" * int(pct / 3) + "░" * (33 - int(pct / 3))
                self._pb_log_write(f"  {sec:<26} {pct:>5.1f}%  {bar}\n", "dim")

            # Projection — three scenarios
            if hasattr(portfolio, "projected_value_at_retirement") and portfolio.projected_value_at_retirement:
                self._pb_log_write(f"\n  {'─'*60}\n", "dim")
                self._pb_log_write(f"  PROJECTED VALUE AT YEAR {years}\n", "header")
                self._pb_log_write(
                    f"  Starting: ${cash:,.0f}  |  "
                    f"Monthly: ${monthly:,.0f}  |  "
                    f"Horizon: {years} years\n", "dim"
                )
                if hasattr(portfolio, "return_was_capped") and portfolio.return_was_capped:
                    self._pb_log_write(
                        f"  ⚠️  Raw portfolio return was {portfolio.expected_return:.1%} — "
                        f"capped at 20% for projection realism.\n"
                        f"  Historical data often overstates future returns.\n", "yellow"
                    )
                self._pb_log_write("\n", "dim")
                for label, val in portfolio.projection_scenarios.items():
                    self._pb_log_write(f"  {label:<40} ${val:>16,.0f}\n", "include")
                self._pb_log_write("\n", "dim")
                self._pb_log_write(
                    "  Note: These are projections, not guarantees. Real returns vary.\n"
                    "  The conservative 7% scenario reflects long-run S&P 500 average.\n", "dim"
                )

            self._pb_log_write("\n")

            # Update pie chart
            self._update_pie(portfolio)
            self._pb_status.config(text=f"Done — {len(portfolio.positions)} positions selected")

        except Exception as e:
            import traceback
            self._pb_log_write(f"\n❌ Error: {e}\n", "red")
            self._pb_log_write(traceback.format_exc(), "dim")
            self._pb_status.config(text="Error")
        finally:
            self._pb_running = False
            self._pb_stop    = False
            self._pb_run_btn.config(state="normal", text="▶  Build Portfolio",
                                     bg=ACCENT, fg="#000000")



    def _draw_empty_pie(self):
        if not self._pb_fig:
            return
        ax = self._pb_ax
        ax.clear()
        ax.set_facecolor(BG)
        ax.pie([1], colors=[BG3], startangle=90)
        ax.set_title("Waiting for portfolio...", color=FG_DIM,
                     fontsize=9, fontfamily="monospace")
        self._pb_canvas.draw()


    def _reload_portfolio(self):
        path = filedialog.askopenfilename(
            title="Select portfolio.xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        self.port_label.config(text="📡 Fetching prices...", fg=FG_DIM)
        def _reload():
            fetch_live_prices(path)
            self.portfolio_ctx = load_portfolio_context(path)
            self.port_label.config(text="📋 Portfolio loaded", fg=GREEN)
        threading.Thread(target=_reload, daemon=True).start()

    # ── ANALYSIS ────────────────────────────

    def _init_prices(self):
        self.port_label.config(text="📡 Fetching prices...", fg=FG_DIM)
        if os.path.exists(PORTFOLIO_FILE):
            fetch_live_prices(PORTFOLIO_FILE)
            self.portfolio_ctx = load_portfolio_context(PORTFOLIO_FILE)
            self.port_label.config(text="📋 Portfolio loaded", fg=GREEN)
        else:
            self.port_label.config(text="📋 No portfolio.xlsx", fg=YELLOW)

    # ── BUILD UI ──────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app = PortfolioBuilderApp(root)
    root.mainloop()
